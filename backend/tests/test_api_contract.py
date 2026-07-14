import pytest
from io import BytesIO
from datetime import datetime, timedelta, timezone


pytest.importorskip("fastapi")
pytest.importorskip("sqlalchemy")

from fastapi.testclient import TestClient  # noqa: E402
from openpyxl import Workbook, load_workbook  # noqa: E402
from sqlalchemy import delete  # noqa: E402

from app.database import SessionLocal  # noqa: E402
from app.main import app, _aggregate_positions  # noqa: E402
from app.models import AIRuntimeConfig, Account, AccountSnapshot, Deal, DecisionCard, InvestorPreference, KlineSnapshot, PositionLayerOverride, PositionSnapshot  # noqa: E402


def test_health_contract():
    with TestClient(app) as client:
        response = client.get("/api/health")
    assert response.status_code == 200
    data = response.json()
    assert "service" in data
    assert "database" in data
    assert "opend" in data
    assert "sqlite_encryption_ready" in data


def test_ai_provider_config_contract_hides_api_key():
    with SessionLocal() as db:
        db.execute(delete(AIRuntimeConfig))
        db.commit()
    with TestClient(app) as client:
        providers = client.get("/api/ai/providers")
        assert providers.status_code == 200
        names = {item["provider"] for item in providers.json()["items"]}
        assert {"deepseek", "openai", "openrouter", "qwen", "custom_openai_compatible"}.issubset(names)

        response = client.put(
            "/api/ai/config",
            json={
                "provider": "openrouter",
                "base_url": "https://openrouter.ai/api/v1",
                "model": "openai/gpt-4o-mini",
                "api_key": "sk-test-secret",
                "enabled": True,
            },
        )
        assert response.status_code == 200
        runtime = response.json()["runtime"]
        assert runtime["provider"] == "openrouter"
        assert runtime["has_api_key"] is True
        assert "sk-test-secret" not in response.text
        assert runtime["masked_api_key"].startswith("sk-")
    with SessionLocal() as db:
        db.execute(delete(AIRuntimeConfig))
        db.commit()


def test_prd_page_contracts_have_empty_states():
    with TestClient(app) as client:
        for path in ("/api/review", "/api/review/trades", "/api/profile", "/api/data/status"):
            response = client.get(path)
            assert response.status_code == 200


def test_excel_import_template_download_contract():
    with TestClient(app) as client:
        response = client.get("/api/import/excel/template")
    assert response.status_code == 200
    assert response.headers["content-type"].startswith("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    workbook = load_workbook(BytesIO(response.content), read_only=True)
    assert workbook.sheetnames == ["账户资产快照", "持仓快照", "成交记录"]


def test_dashboard_includes_freshness_matrix(monkeypatch):
    import app.services.fx_rates as fx_rates

    monkeypatch.setattr(fx_rates, "fetch_frankfurter_rate", lambda base, quote: {"rate": 7.8 if quote == "USD" else 1.1, "date": "2026-07-08"})
    with TestClient(app) as client:
        response = client.get("/api/dashboard")
        assert response.status_code == 200
        data = response.json()
        assert "freshness" in data
        assert "base_currency" in data["portfolio"]
        assert "display_rate_meta" in data["portfolio"]
        assert data["portfolio"]["display_rates"][data["portfolio"]["base_currency"]] == 1


def test_position_contract_includes_market_value_currency():
    with TestClient(app) as client:
        response = client.get("/api/positions")
        assert response.status_code == 200
        data = response.json()
        for item in data["items"]:
            assert "raw_currency" in item
            assert "normalized_currency" in item
            assert "exchange_rate_to_base" in item


def test_aggregate_positions_uses_portfolio_weight_for_all_accounts():
    hk_account = "__weight_hk_account__"
    us_account = "__weight_us_account__"
    snapshot_time = datetime(2026, 7, 4, 10, 0, tzinfo=timezone.utc)
    positions = [
        PositionSnapshot(
            account_id=hk_account,
            code="US.QQQ",
            name="QQQ",
            market="US",
            asset_type="fund",
            quantity=1,
            average_cost=100,
            current_price=100,
            raw_market_value=8000,
            raw_currency="HKD",
            normalized_market_value=8000,
            normalized_currency="HKD",
            position_weight=0.8,
            profit_loss_ratio=0,
            position_layer="核心长期仓",
            layer_source="system",
            layer_confidence="高",
            layer_reason="test",
            snapshot_time=snapshot_time,
            sync_id="weight_hk",
        ),
        PositionSnapshot(
            account_id=hk_account,
            code="US.NVDA",
            name="NVDA",
            market="US",
            asset_type="stock",
            quantity=1,
            average_cost=100,
            current_price=100,
            raw_market_value=1000,
            raw_currency="HKD",
            normalized_market_value=1000,
            normalized_currency="HKD",
            position_weight=0.1,
            profit_loss_ratio=0.1,
            position_layer="中期配置仓",
            layer_source="system",
            layer_confidence="高",
            layer_reason="test",
            snapshot_time=snapshot_time,
            sync_id="weight_hk",
        ),
        PositionSnapshot(
            account_id=us_account,
            code="US.NVDA",
            name="NVDA",
            market="US",
            asset_type="stock",
            quantity=1,
            average_cost=100,
            current_price=100,
            raw_market_value=80,
            raw_currency="USD",
            normalized_market_value=80,
            normalized_currency="USD",
            position_weight=0.8,
            profit_loss_ratio=0.2,
            position_layer="中期配置仓",
            layer_source="system",
            layer_confidence="高",
            layer_reason="test",
            snapshot_time=snapshot_time,
            sync_id="weight_us",
        ),
    ]

    total_assets_cny = 10000 * 0.92 + 100 * 7.2
    qqq_weight = 8000 * 0.92 / total_assets_cny
    nvda_weight = (1000 * 0.92 + 80 * 7.2) / total_assets_cny
    aggregated = {item["code"]: item for item in _aggregate_positions(positions, total_assets_cny, "CNY")}

    assert aggregated["US.QQQ"]["position_weight"] == pytest.approx(qqq_weight)
    assert aggregated["US.NVDA"]["position_weight"] == pytest.approx(nvda_weight)
    assert aggregated["US.NVDA"]["position_weight"] < 0.2


def test_all_account_position_detail_with_decision_card():
    code = "__DETAIL.TEST__"
    account_ids = ["__detail_account_a__", "__detail_account_b__"]
    snapshot_time = datetime(2026, 7, 4, 10, 0)
    try:
        for account_id in account_ids:
            _create_account(account_id)
        with SessionLocal() as db:
            for index, account_id in enumerate(account_ids):
                db.add(AccountSnapshot(account_id=account_id, total_assets=1000, cash=500, market_value=500, raw_currency_values={"currency": "CNY"}, snapshot_time=snapshot_time, sync_id=f"detail_asset_{index}"))
                db.add(
                    PositionSnapshot(
                        account_id=account_id,
                        code=code,
                        name="详情测试",
                        market="US",
                        asset_type="stock",
                        quantity=1,
                        average_cost=100,
                        current_price=100,
                        raw_market_value=100,
                        raw_currency="CNY",
                        normalized_market_value=100,
                        normalized_currency="CNY",
                        position_weight=0.1,
                        profit_loss_ratio=0,
                        position_layer="中期配置仓",
                        layer_source="system",
                        layer_confidence="高",
                        layer_reason="test",
                        snapshot_time=snapshot_time,
                        sync_id=f"detail_position_{index}",
                    )
                )
            db.add(
                DecisionCard(
                    card_id="detail_card_test",
                    code=code,
                    position_layer="中期配置仓",
                    recommendation="观察",
                    confidence="中",
                    reasons=["test"],
                    risks=[],
                    key_prices={},
                    data_time=snapshot_time - timedelta(days=1),
                    data_version="test",
                    status="正常",
                    priority="P3",
                    created_at=snapshot_time,
                )
            )
            db.commit()

        with TestClient(app) as client:
            response = client.get(f"/api/positions/{code}", params={"account_id": "all"})

        assert response.status_code == 200
        data = response.json()
        assert data["position"]["account_id"] == "all"
        assert len(data["account_positions"]) == 2
        assert data["cards"][0]["needs_regeneration"] is True
    finally:
        with SessionLocal() as db:
            db.execute(delete(DecisionCard).where(DecisionCard.code == code))
            db.commit()
        for account_id in account_ids:
            _cleanup_account(account_id)


def test_position_kline_contract_uses_local_snapshot():
    account_id = "__kline_contract_account__"
    code = "US.QQQ"
    snapshot_time = datetime(2026, 7, 4, 10, 0)

    try:
        _create_account(account_id)
        with SessionLocal() as db:
            db.execute(delete(KlineSnapshot).where(KlineSnapshot.code == code, KlineSnapshot.period == "K_DAY", KlineSnapshot.snapshot_time == snapshot_time))
            db.add(
                PositionSnapshot(
                    account_id=account_id,
                    code=code,
                    name="QQQ",
                    market="US",
                    asset_type="fund",
                    quantity=1,
                    average_cost=100,
                    current_price=101,
                    raw_market_value=101,
                    raw_currency="USD",
                    normalized_market_value=101,
                    normalized_currency="USD",
                    position_weight=0.1,
                    profit_loss_ratio=0.01,
                    position_layer="核心长期仓",
                    layer_source="system",
                    layer_confidence="高",
                    layer_reason="test",
                    snapshot_time=snapshot_time,
                    sync_id="kline_contract_position",
                )
            )
            db.add_all([
                KlineSnapshot(code=code, provider="futu", period="K_DAY", time_key="2026-07-01", open=100, close=101, high=102, low=99, volume=1200, turnover=121000, snapshot_time=snapshot_time, sync_id="kline_contract"),
                KlineSnapshot(code=code, provider="futu", period="K_DAY", time_key="2026-07-02", open=101, close=100, high=103, low=98, volume=900, turnover=90000, snapshot_time=snapshot_time, sync_id="kline_contract"),
            ])
            db.commit()

        with TestClient(app) as client:
            response = client.get(f"/api/positions/kline/{code}", params={"account_id": account_id, "ktype": "K_DAY", "count": 30})

        assert response.status_code == 200
        data = response.json()
        assert data["code"] == code
        assert data["status"] == "available"
        assert data["items"][0]["close"] == 101
    finally:
        with SessionLocal() as db:
            db.execute(delete(KlineSnapshot).where(KlineSnapshot.code == code, KlineSnapshot.period == "K_DAY", KlineSnapshot.snapshot_time == snapshot_time))
            db.commit()
        _cleanup_account(account_id)


def test_account_position_snapshot_can_be_created_and_edited():
    account_id = "__manual_position_account__"
    try:
        _create_account(account_id)
        snapshot_time = datetime(2026, 7, 4, 10, 0, tzinfo=timezone.utc)
        with SessionLocal() as db:
            db.add(AccountSnapshot(account_id=account_id, total_assets=10000, cash=5000, market_value=5000, raw_currency_values={"currency": "CNY"}, snapshot_time=snapshot_time, sync_id="manual_position_asset"))
            db.commit()

        payload = {
            "code": "CN.TEST",
            "name": "手动持仓",
            "market": "CN",
            "asset_type": "stock",
            "quantity": 100,
            "average_cost": 8,
            "current_price": 10,
            "market_value": 1000,
            "currency": "CNY",
            "normalized_market_value": 1000,
            "normalized_currency": "CNY",
            "profit_loss_ratio": 0.25,
            "position_layer": "中期配置仓",
            "snapshot_time": "2026-07-04T10:00:00",
        }
        with TestClient(app) as client:
            response = client.post(f"/api/data/accounts/{account_id}/positions", json=payload)
            assert response.status_code == 200
            item = response.json()["overview"]["positions"][0]
            assert item["code"] == "CN.TEST"
            assert item["position_weight"] == 0.1

            response = client.post(
                f"/api/data/accounts/{account_id}/positions",
                json={**payload, "original_code": "CN.TEST", "original_snapshot_time": "2026-07-04T10:00:00", "name": "已编辑持仓", "current_price": 12, "market_value": 1200, "normalized_market_value": 1200},
            )
            assert response.status_code == 200
            positions = response.json()["overview"]["positions"]
            assert len(positions) == 1
            assert positions[0]["name"] == "已编辑持仓"
            assert positions[0]["position_weight"] == 0.12
    finally:
        _cleanup_account(account_id)


def test_account_position_snapshot_accepts_minimal_calculated_payload():
    account_id = "__manual_position_minimal_account__"
    try:
        _create_account(account_id)
        snapshot_time = datetime(2026, 7, 4, 10, 0, tzinfo=timezone.utc)
        with SessionLocal() as db:
            db.add(AccountSnapshot(account_id=account_id, total_assets=10000, cash=9000, market_value=1000, raw_currency_values={"currency": "CNY"}, snapshot_time=snapshot_time, sync_id="manual_position_minimal_asset"))
            db.commit()

        payload = {
            "code": "US.MIN",
            "quantity": 5,
            "current_price": 20,
            "snapshot_time": "2026-07-04T10:00:00",
        }
        with TestClient(app) as client:
            response = client.post(f"/api/data/accounts/{account_id}/positions", json=payload)
            assert response.status_code == 200
            item = response.json()["overview"]["positions"][0]
            assert item["name"] == "US.MIN"
            assert item["raw_market_value"] == 100
            assert item["normalized_market_value"] == 100
            assert item["profit_loss_ratio"] == 0
            assert item["position_weight"] == 0.01

            response = client.post(
                f"/api/data/accounts/{account_id}/positions",
                json={
                    **payload,
                    "original_code": "US.MIN",
                    "original_snapshot_time": "2026-07-04T10:00:00",
                    "code": "US.MIN",
                    "currency": "USD",
                    "normalized_currency": "CNY",
                    "exchange_rate_to_base": 7.2,
                },
            )
            assert response.status_code == 200
            item = response.json()["overview"]["positions"][0]
            assert item["raw_market_value"] == 100
            assert item["normalized_market_value"] == 720
            assert item["position_weight"] == pytest.approx(0.072)
    finally:
        _cleanup_account(account_id)


def test_account_position_snapshot_can_be_deleted():
    account_id = "__manual_position_delete_account__"
    snapshot_time = datetime(2026, 7, 4, 10, 0)
    try:
        _create_account(account_id)
        with SessionLocal() as db:
            db.add(AccountSnapshot(account_id=account_id, total_assets=1000, cash=0, market_value=1000, raw_currency_values={"currency": "CNY"}, snapshot_time=snapshot_time, sync_id="manual_position_delete_asset"))
            db.add(
                PositionSnapshot(
                    account_id=account_id,
                    code="CN.DELETE",
                    name="待删除持仓",
                    market="CN",
                    asset_type="stock",
                    quantity=100,
                    average_cost=10,
                    current_price=10,
                    raw_market_value=1000,
                    raw_currency="CNY",
                    normalized_market_value=1000,
                    normalized_currency="CNY",
                    position_weight=1,
                    profit_loss_ratio=0,
                    position_layer="中期配置仓",
                    layer_source="user",
                    layer_confidence="高",
                    layer_reason="用户手动维护持仓快照",
                    snapshot_time=snapshot_time,
                    sync_id="manual_position_delete",
                )
            )
            db.commit()

        with TestClient(app) as client:
            response = client.delete(f"/api/data/accounts/{account_id}/positions?code=CN.DELETE&snapshot_time={snapshot_time.isoformat()}")
            assert response.status_code == 200
            assert response.json()["overview"]["positions"] == []

            response = client.delete(f"/api/data/accounts/{account_id}/positions?code=CN.DELETE&snapshot_time={snapshot_time.isoformat()}")
            assert response.status_code == 404
    finally:
        _cleanup_account(account_id)


def test_account_deal_can_be_created_edited_and_deleted():
    account_id = "__manual_deal_account__"
    try:
        _create_account(account_id)
        payload = {
            "deal_id": "d_manual_1",
            "order_id": "o_manual_1",
            "code": "US.AAPL",
            "side": "BUY",
            "price": 180,
            "quantity": 2,
            "deal_time": "2026-07-04T10:00:00",
            "market": "US",
        }
        with TestClient(app) as client:
            response = client.post(f"/api/data/accounts/{account_id}/deals", json=payload)
            assert response.status_code == 200
            deals = response.json()["overview"]["deals"]
            assert len(deals) == 1
            assert deals[0]["deal_id"] == "d_manual_1"
            assert deals[0]["code"] == "US.AAPL"

            response = client.post(
                f"/api/data/accounts/{account_id}/deals",
                json={**payload, "original_deal_id": "d_manual_1", "deal_id": "d_manual_2", "code": "US.MSFT", "price": 210},
            )
            assert response.status_code == 200
            deals = response.json()["overview"]["deals"]
            assert len(deals) == 1
            assert deals[0]["deal_id"] == "d_manual_2"
            assert deals[0]["code"] == "US.MSFT"

            response = client.delete(f"/api/data/accounts/{account_id}/deals/d_manual_2")
            assert response.status_code == 200
            assert response.json()["overview"]["deals"] == []

            response = client.delete(f"/api/data/accounts/{account_id}/deals/d_manual_2")
            assert response.status_code == 404
    finally:
        _cleanup_account(account_id)


def test_account_position_edit_matches_timezone_shifted_original_time():
    account_id = "__manual_position_timezone_account__"
    try:
        _create_account(account_id)
        snapshot_time = datetime(2026, 7, 9, 5, 42)
        with SessionLocal() as db:
            db.add(AccountSnapshot(account_id=account_id, total_assets=8266, cash=1011, market_value=7255, raw_currency_values={"currency": "USD"}, snapshot_time=snapshot_time, sync_id="manual_position_asset_tz"))
            db.add(
                PositionSnapshot(
                    account_id=account_id,
                    code="US.NVDA",
                    name="英伟达",
                    market="US",
                    asset_type="stock",
                    quantity=28,
                    average_cost=207.34,
                    current_price=198,
                    raw_market_value=55424,
                    raw_currency="USD",
                    normalized_market_value=55424,
                    normalized_currency="USD",
                    position_weight=6.705,
                    profit_loss_ratio=-0.045,
                    position_layer="核心长期仓",
                    layer_source="user",
                    layer_confidence="高",
                    layer_reason="用户手动维护持仓快照",
                    snapshot_time=snapshot_time,
                    sync_id="manual_position_tz",
                )
            )
            db.commit()

        payload = {
            "original_code": "US.NVDA",
            "original_snapshot_time": (snapshot_time - timedelta(hours=8)).replace(tzinfo=timezone.utc).isoformat(),
            "code": "US.NVDA",
            "name": "英伟达",
            "market": "US",
            "asset_type": "stock",
            "quantity": 28,
            "average_cost": 207.34,
            "current_price": 198,
            "market_value": 5544.24,
            "currency": "USD",
            "normalized_market_value": 5544.24,
            "normalized_currency": "USD",
            "profit_loss_ratio": -0.045,
            "position_layer": "核心长期仓",
            "snapshot_time": snapshot_time.isoformat(),
        }
        with TestClient(app) as client:
            response = client.post(f"/api/data/accounts/{account_id}/positions", json=payload)
            assert response.status_code == 200
            positions = response.json()["overview"]["positions"]
            assert len(positions) == 1
            assert positions[0]["normalized_market_value"] == 5544.24
            assert positions[0]["position_weight"] == pytest.approx(5544.24 / 8266)

        with SessionLocal() as db:
            saved = db.query(PositionSnapshot).filter(PositionSnapshot.account_id == account_id, PositionSnapshot.code == "US.NVDA").all()
            assert len(saved) == 1
            assert saved[0].normalized_market_value == 5544.24
    finally:
        _cleanup_account(account_id)


def test_position_layer_override_updates_latest_snapshot_immediately():
    account_id = "__layer_override_account__"
    code = "__LAYER_OVERRIDE__"
    try:
        _create_account(account_id)
        snapshot_time = datetime(2026, 7, 4, 10, 0, tzinfo=timezone.utc)
        with SessionLocal() as db:
            db.add(
                PositionSnapshot(
                    account_id=account_id,
                    code=code,
                    name="仓位类型测试",
                    market="CN",
                    asset_type="fund",
                    quantity=100,
                    average_cost=1,
                    current_price=1,
                    raw_market_value=100,
                    raw_currency="CNY",
                    normalized_market_value=100,
                    normalized_currency="CNY",
                    position_weight=1,
                    profit_loss_ratio=0,
                    position_layer="中期配置仓",
                    layer_source="system",
                    layer_confidence="中",
                    layer_reason="系统识别",
                    snapshot_time=snapshot_time,
                    sync_id="contract_layer_override",
                )
            )
            db.commit()

        with TestClient(app) as client:
            response = client.patch(
                f"/api/positions/{code}/layer",
                json={"position_layer": "遗留观察仓", "reason": "用户在工作台手动修正"},
            )
            assert response.status_code == 200
            assert response.json()["updated_positions"] == 1

            response = client.get(f"/api/positions?account_id={account_id}")
            assert response.status_code == 200
            item = response.json()["items"][0]
            assert item["position_layer"] == "遗留观察仓"
            assert item["layer_source"] == "user"
            assert item["layer_confidence"] == "高"

        with SessionLocal() as db:
            saved = db.get(PositionLayerOverride, code)
            latest = db.query(PositionSnapshot).filter(PositionSnapshot.account_id == account_id, PositionSnapshot.code == code).one()
            assert saved is not None
            assert saved.position_layer == "遗留观察仓"
            assert latest.position_layer == "遗留观察仓"
    finally:
        with SessionLocal() as db:
            db.execute(delete(PositionLayerOverride).where(PositionLayerOverride.code == code))
            db.commit()
        _cleanup_account(account_id)


def test_position_layer_override_is_applied_when_listing_positions():
    account_id = "__layer_override_read_account__"
    code = "__LAYER_OVERRIDE_READ__"
    try:
        _create_account(account_id)
        snapshot_time = datetime(2026, 7, 4, 10, 0, tzinfo=timezone.utc)
        with SessionLocal() as db:
            db.add(
                PositionSnapshot(
                    account_id=account_id,
                    code=code,
                    name="仓位读取测试",
                    market="CN",
                    asset_type="fund",
                    quantity=100,
                    average_cost=1,
                    current_price=1,
                    raw_market_value=100,
                    raw_currency="CNY",
                    normalized_market_value=100,
                    normalized_currency="CNY",
                    position_weight=1,
                    profit_loss_ratio=0,
                    position_layer="中期配置仓",
                    layer_source="system",
                    layer_confidence="中",
                    layer_reason="系统识别",
                    snapshot_time=snapshot_time,
                    sync_id="contract_layer_override_read",
                )
            )
            db.add(
                PositionLayerOverride(
                    code=code,
                    position_layer="核心长期仓",
                    reason="用户在工作台手动修正",
                    updated_at=snapshot_time,
                )
            )
            db.commit()

        with TestClient(app) as client:
            response = client.get(f"/api/positions?account_id={account_id}")
            assert response.status_code == 200
            item = response.json()["items"][0]
            assert item["position_layer"] == "核心长期仓"
            assert item["layer_source"] == "user"
            assert item["layer_reason"] == "用户在工作台手动修正"
    finally:
        with SessionLocal() as db:
            db.execute(delete(PositionLayerOverride).where(PositionLayerOverride.code == code))
            db.commit()
        _cleanup_account(account_id)


def test_profile_preferences_round_trip():
    test_account_id = "__contract_test_preferences__"
    payload = {
        "kyc_profile": {"summary": "长期资金"},
        "risk_tolerance": "稳健",
        "investment_horizon": "1-3年",
        "liquidity_needs": "保留6个月现金",
        "target_return": "年化8%-12%",
        "notes": "测试偏好",
    }
    try:
        with TestClient(app) as client:
            response = client.patch(f"/api/profile/preferences?account_id={test_account_id}", json=payload)
            assert response.status_code == 200
            saved = response.json()
            assert saved["account_id"] == test_account_id
            assert saved["risk_tolerance"] == "稳健"

            response = client.get(f"/api/profile/preferences?account_id={test_account_id}")
            assert response.status_code == 200
            assert response.json()["target_return"] == "年化8%-12%"
    finally:
        with SessionLocal() as db:
            db.execute(delete(InvestorPreference).where(InvestorPreference.account_id == test_account_id))
            db.commit()


def test_profile_workflow_stream_contract():
    with TestClient(app) as client:
        response = client.post("/api/profile/ai-workflows/customer_profile", json={"consent_external_ai": True, "use_external_model": False})
        assert response.status_code == 200
        run_id = response.json()["run"]["run_id"]

        with client.stream("GET", f"/api/profile/ai-workflows/{run_id}/stream") as stream:
            body = "".join(stream.iter_text())
        assert "event: run_started" in body
        assert "event: agent_warning" in body
        assert "event: run_failed" in body
        assert "没有大模型参与时不生成投顾报告" in body


def test_profile_workflow_rejects_unknown_type():
    with TestClient(app) as client:
        response = client.post("/api/profile/ai-workflows/unknown", json={"consent_external_ai": True})
    assert response.status_code == 400


def test_account_crud_and_data_overview_contract():
    account_id = "__contract_account__"
    try:
        with TestClient(app) as client:
            response = client.post(
                "/api/accounts",
                json={
                    "display_name": "测试账户",
                    "import_modes": ["api", "local"],
                    "position_import_modes": ["api"],
                    "review_import_modes": ["local"],
                    "market_data_provider": "tushare",
                    "news_data_provider": "futu",
                    "base_currency": "CNY",
                    "markets": ["CN"],
                    "enabled": True,
                },
            )
            assert response.status_code == 200
            account_id = response.json()["account"]["account_id"]
            assert account_id.startswith("acct_")
            assert response.json()["account"]["import_modes"] == ["api", "local"]
            assert response.json()["account"]["position_import_modes"] == ["api"]
            assert response.json()["account"]["review_import_modes"] == ["local"]
            assert response.json()["account"]["market_data_provider"] == "tushare"
            assert response.json()["account"]["news_data_provider"] == "futu"

            response = client.patch(
                f"/api/accounts/{account_id}",
                json={
                    "enabled": False,
                    "account_type": "证券账户",
                    "import_modes": ["local"],
                    "position_import_modes": ["local"],
                    "review_import_modes": ["api", "local"],
                    "market_data_provider": "akshare",
                    "news_data_provider": "marketaux",
                },
            )
            assert response.status_code == 200
            assert response.json()["account"]["enabled"] is False
            assert response.json()["account"]["import_mode"] == "local"
            assert response.json()["account"]["import_modes"] == ["local"]
            assert response.json()["account"]["position_import_modes"] == ["local"]
            assert response.json()["account"]["review_import_modes"] == ["api", "local"]
            assert response.json()["account"]["market_data_provider"] == "akshare"
            assert response.json()["account"]["news_data_provider"] == "marketaux"
            with SessionLocal() as db:
                saved_account = db.get(Account, account_id)
                assert saved_account is not None
                assert saved_account.position_import_modes == "local"
                assert saved_account.review_import_modes == "api,local"
                assert saved_account.market_data_provider == "akshare"
                assert saved_account.news_data_provider == "marketaux"

            response = client.get(f"/api/data/accounts/{account_id}/overview")
            assert response.status_code == 200
            data = response.json()
            assert data["account"]["account_id"] == account_id
            assert data["account"]["import_modes"] == ["local"]
            assert data["account"]["position_import_modes"] == ["local"]
            assert data["account"]["review_import_modes"] == ["api", "local"]
            assert data["account"]["market_data_provider"] == "akshare"
            quote_states = [item for item in data["provider_states"] if item["data_type"] == "quote"]
            assert quote_states
            assert {item["provider"] for item in quote_states} == {"akshare"}
            assert data["asset_snapshot"] is None
            assert data["positions"] == []
            assert data["deals"] == []

            response = client.delete(f"/api/accounts/{account_id}")
            assert response.status_code == 200
    finally:
        _cleanup_account(account_id)


def test_excel_preview_validates_required_fields_and_imports_rows():
    account_id = "__excel_account__"
    try:
        _create_account(account_id)
        workbook = Workbook()
        asset = workbook.active
        asset.title = "账户资产快照"
        asset.append(["snapshot_time", "total_assets", "cash", "market_value", "currency"])
        asset.append(["2026-07-04 10:00:00", 10000, 1000, 9000, "CNY"])
        position = workbook.create_sheet("持仓快照")
        position.append(["snapshot_time", "code", "name", "market", "asset_type", "quantity", "current_price", "market_value", "currency", "average_cost"])
        position.append(["2026-07-04 10:00:00", "US.NVDA", "英伟达", "US", "stock", 10, 120, 1200, "USD", 100])
        deal = workbook.create_sheet("成交记录")
        deal.append(["deal_id", "code", "side", "price", "quantity", "deal_time"])
        deal.append(["d1", "US.NVDA", "BUY", 100, 10, "2026-07-01 10:00:00"])
        payload = BytesIO()
        workbook.save(payload)

        with TestClient(app) as client:
            response = client.post(
                f"/api/import/excel/preview?account_id={account_id}",
                files={"file": ("import.xlsx", payload.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            )
            assert response.status_code == 200
            preview = response.json()
            assert preview["errors"] == []
            assert preview["position_count"] == 1

            response = client.post("/api/import/excel/confirm", json=preview)
            assert response.status_code == 200
            assert response.json()["status"] == "成功"

            overview = client.get(f"/api/data/accounts/{account_id}/overview").json()
            assert overview["asset_snapshot"]["total_assets"] == 10000
            assert overview["positions"][0]["code"] == "US.NVDA"
            assert overview["deals"][0]["deal_id"] == "d1"
    finally:
        _cleanup_account(account_id)


def test_file_preview_reports_unavailable_ocr_for_images(monkeypatch):
    from app.services import imports as import_service
    from app.services.ocr import DisabledOCRProvider

    monkeypatch.setattr(import_service, "get_ocr_provider", lambda: DisabledOCRProvider())
    account_id = "__ocr_disabled__"
    try:
        _create_account(account_id)
        with TestClient(app) as client:
            response = client.post(
                f"/api/import/file/preview?account_id={account_id}",
                files={"file": ("screen.png", b"not-a-real-image", "image/png")},
            )
        assert response.status_code == 200
        data = response.json()
        assert data["can_confirm"] is False
        assert "本地 OCR 未启用" in data["errors"][0]
    finally:
        _cleanup_account(account_id)


def test_alipay_pdf_never_falls_back_to_sample_data():
    with TestClient(app) as client:
        response = client.post(
            "/api/import/alipay/preview",
            files={"file": ("bad.pdf", b"not a real alipay pdf", "application/pdf")},
        )
    assert response.status_code == 400
    assert "样例" not in response.text
    assert "sample_template" not in response.text


def _create_account(account_id: str) -> None:
    with SessionLocal() as db:
        db.add(Account(account_id=account_id, source_name="manual", display_name=account_id, import_mode="manual", base_currency="CNY"))
        db.commit()


def _cleanup_account(account_id: str) -> None:
    with SessionLocal() as db:
        db.execute(delete(Deal).where(Deal.account_id == account_id))
        db.execute(delete(PositionSnapshot).where(PositionSnapshot.account_id == account_id))
        db.execute(delete(AccountSnapshot).where(AccountSnapshot.account_id == account_id))
        db.execute(delete(Account).where(Account.account_id == account_id))
        db.commit()
