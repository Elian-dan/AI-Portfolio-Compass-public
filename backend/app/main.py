from __future__ import annotations

from datetime import datetime, timedelta, timezone
from io import BytesIO
import json
import socket
from xml.sax.saxutils import escape
from zipfile import ZIP_DEFLATED, ZipFile
from typing import Optional
from uuid import uuid4

from fastapi import Body, Depends, FastAPI, File, HTTPException, Query, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from sqlalchemy import delete, func, select, text
from sqlalchemy.orm import Session

from app.adapters.futu_adapter import FutuReadOnlyAdapter
from app.config import get_settings
from app.database import engine, get_db, init_db
from app.models import AIAnalysis, Account, AccountSnapshot, AIWorkflowRun, Alert, DataSourceState, Deal, DecisionCard, KlineSnapshot, NewsItem, PositionLayerOverride, PositionSnapshot, QuoteSummary, TradeReview, UserAction
from app.schemas import DealSaveRequest, DeleteLocalDataRequest, LayerOverrideRequest, PositionSnapshotSaveRequest
from app.services.sync import (
    delete_local_data,
    latest_account_snapshots,
    latest_cards,
    latest_positions,
    latest_profile,
    latest_review,
    latest_source_states,
    latest_sync,
    mark_source_state,
    pull_account_market_data,
    recent_news_for_code,
    freshness_summary,
    run_manual_sync,
)
from app.services.ai_config import get_ai_scene_config, list_ai_scene_configs, save_ai_scene_configs, scene_key
from app.services.ai_decision_cards import SYSTEM_INSTRUCTION as DECISION_CARD_SYSTEM_INSTRUCTION
from app.services.ai_decision_cards import generate_ai_decision_cards
from app.services.ai_engine import (
    SYSTEM_INSTRUCTION as AI_ANALYSIS_SYSTEM_INSTRUCTION,
    call_llm_payload,
    generate_ai_analysis,
    latest_ai_analysis,
    normalize_ai_output,
)
from app.services.ai_runtime import (
    active_ai_runtime,
    ai_runtime_config_payload,
    mark_ai_runtime_test,
    provider_templates_payload,
    public_runtime,
    runtime_can_call_external,
    save_ai_runtime_config,
)
from app.services.imports import confirm_import, preview_import
from app.services.fx_rates import display_rates as fx_display_rates
from app.services.profile_workflows import (
    SYSTEM_INSTRUCTION as WORKFLOW_SYSTEM_INSTRUCTION,
    WORKFLOW_LABELS,
    WORKFLOW_TYPES,
    cancel_workflow_run,
    create_workflow_run,
    delete_workflow_run,
    get_investor_preference,
    latest_workflow_runs,
    preference_to_dict,
    stream_workflow_run,
    upsert_investor_preference,
    workflow_markdown_filename,
    workflow_run_to_dict,
)
from app.services.trade_review import trade_review_payload, update_trade_review_intent
from app.services.providers import account_provider_statuses, normalize_market, provider_registry


app = FastAPI(title="AI Portfolio Compass", version="0.1.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=[],
    allow_origin_regex=r"^http://(localhost|127\.0\.0\.1):\d+$",
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.on_event("startup")
def on_startup() -> None:
    init_db()


@app.get("/api/health")
def health(db: Session = Depends(get_db)) -> dict:
    settings = get_settings()
    demo_mode = settings.workspace == "demo" or (
        settings.database_url.startswith("sqlite:///") and "demo" in settings.database_url.lower()
    )
    db_status = "ok"
    try:
        with engine.connect() as conn:
            conn.execute(text("select 1"))
    except Exception as exc:
        db_status = f"error: {str(exc)[:120]}"

    connected, opend_status = (False, "演示工作区不连接真实账户") if demo_mode else _opend_health_with_timeout()
    futu_accounts = 0
    futu_account_access = False
    futu_account_message = "OpenD 未连接"
    if connected:
        try:
            futu_account_access, futu_accounts, futu_account_message = FutuReadOnlyAdapter().account_access()
        except Exception as exc:  # pragma: no cover - depends on local OpenD
            futu_account_message = str(exc)[:160]
    local_counts = {
        "accounts": db.scalar(select(func.count()).select_from(Account)) or 0,
        "account_snapshots": db.scalar(select(func.count()).select_from(AccountSnapshot)) or 0,
        "positions": db.scalar(select(func.count()).select_from(PositionSnapshot)) or 0,
        "deals": db.scalar(select(func.count()).select_from(Deal)) or 0,
    }
    runtime = active_ai_runtime(db)
    return {
        "service": "ok",
        "database": db_status,
        "opend": "connected" if connected else opend_status,
        "futu": {
            "host": settings.futu_host,
            "port": settings.futu_port,
            "opend_connected": connected,
            "account_access": futu_account_access,
            "account_count": futu_accounts,
            "message": futu_account_message,
        },
        "local_data": {
            **local_counts,
            "empty": not any(local_counts.values()),
        },
        "demo_mode": demo_mode,
        "workspace": "demo" if demo_mode else "formal",
        "sqlite_encryption_ready": False,
        "ai": {
            "provider": runtime["provider"],
            "display_name": runtime["display_name"],
            "model": runtime["model"] if runtime_can_call_external(runtime) else "local_reasoning",
            "configured": runtime_can_call_external(runtime),
            "enabled": bool(runtime.get("enabled")),
            "source": runtime.get("source", ""),
        },
    }


@app.post("/api/sync/manual")
def manual_sync(db: Session = Depends(get_db)) -> dict:
    settings = get_settings()
    if settings.workspace == "demo" or "demo" in settings.database_url.lower():
        return {
            "sync_id": "demo_workspace",
            "status": "已跳过",
            "message": "演示工作区不会同步真实账户；请切换到正式模式后再同步。",
        }
    task = run_manual_sync(db)
    return {
        "sync_id": task.sync_id,
        "status": task.status,
        "message": task.error_message or "同步完成",
    }


@app.get("/api/ai/providers")
def ai_providers() -> dict:
    return {"items": provider_templates_payload()}


@app.get("/api/ai/config")
def ai_config(db: Session = Depends(get_db)) -> dict:
    return _ai_config_payload(db)


@app.put("/api/ai/config")
def save_ai_config(payload: dict = Body(default_factory=dict), db: Session = Depends(get_db)) -> dict:
    save_ai_runtime_config(db, payload)
    scenes = payload.get("scenes")
    if isinstance(scenes, dict):
        scenes = [{"scene": key, **value} for key, value in scenes.items() if isinstance(value, dict)]
    if isinstance(scenes, list):
        defaults = _ai_scene_defaults(db)
        save_ai_scene_configs(db, scenes, {item["scene"] for item in defaults})
    return _ai_config_payload(db)


@app.post("/api/ai/config/test")
def test_ai_config(payload: dict = Body(default_factory=dict), db: Session = Depends(get_db)) -> dict:
    runtime = active_ai_runtime(db)
    if payload:
        runtime = {
            **runtime,
            "provider": str(payload.get("provider") or runtime.get("provider") or ""),
            "display_name": str(payload.get("display_name") or runtime.get("display_name") or ""),
            "base_url": str(payload.get("base_url") or runtime.get("base_url") or "").rstrip("/"),
            "model": str(payload.get("model") or runtime.get("model") or ""),
            "enabled": bool(payload.get("enabled", runtime.get("enabled", True))),
            "api_key": str(payload.get("api_key") or runtime.get("api_key") or ""),
        }
        runtime["has_api_key"] = bool(str(runtime.get("api_key") or "").strip())
    if not runtime_can_call_external(runtime):
        message = "请先填写 API Key、Base URL 和模型名，并启用外部大模型。"
        mark_ai_runtime_test(db, "failed", message)
        return {"status": "failed", "message": message, "runtime": public_runtime(runtime)}
    test_payload = {
        "model": runtime["model"],
        "messages": [
            {"role": "system", "content": "You are a connection test endpoint. Reply with strict JSON."},
            {"role": "user", "content": "Return {\"status\":\"ok\"} only."},
        ],
        "temperature": 0,
        "response_format": {"type": "json_object"},
    }
    try:
        output = call_llm_payload(test_payload, runtime["api_key"], runtime["base_url"], timeout=20)
        message = "连接可用"
        mark_ai_runtime_test(db, "success", message)
        return {"status": "success", "message": message, "output": output, "runtime": public_runtime(runtime)}
    except Exception as exc:
        message = str(exc)[:500]
        mark_ai_runtime_test(db, "failed", message)
        return {"status": "failed", "message": message, "runtime": public_runtime(runtime)}


def _ai_scene_defaults(db: Session | None = None) -> list[dict[str, str]]:
    settings = get_settings()
    runtime = active_ai_runtime(db) if db else None
    default_model = str(runtime.get("model") if runtime else settings.deepseek_model) or settings.deepseek_model
    default_provider = str(runtime.get("provider") if runtime else settings.ai_provider) or "deepseek"
    workflow_defaults = [
        {
            "scene": scene_key(workflow_type),
            "label": {
                "portfolio_diagnosis": "组合诊断报告",
                "customer_profile": "客户画像报告",
                "asset_allocation": "资产配置报告",
            }[workflow_type],
            "provider": default_provider,
            "model": default_model,
            "system_prompt": WORKFLOW_SYSTEM_INSTRUCTION,
        }
        for workflow_type in ("portfolio_diagnosis", "customer_profile", "asset_allocation")
    ]
    return [
        {
            "scene": "planning_agent",
            "label": "规划 Agent",
            "provider": default_provider,
            "model": default_model,
            "system_prompt": "",
        },
        *workflow_defaults,
        {
            "scene": "ai_analysis",
            "label": "单标的 AI 分析",
            "provider": default_provider,
            "model": default_model,
            "system_prompt": AI_ANALYSIS_SYSTEM_INSTRUCTION,
        },
        {
            "scene": "decision_cards",
            "label": "持仓诊断卡",
            "provider": default_provider,
            "model": default_model,
            "system_prompt": DECISION_CARD_SYSTEM_INSTRUCTION,
        },
    ]


def _ai_config_payload(db: Session) -> dict:
    defaults = _ai_scene_defaults(db)
    scenes = list_ai_scene_configs(db, defaults)
    return ai_runtime_config_payload(db, scenes)


@app.get("/api/sync/status")
def sync_status(db: Session = Depends(get_db)) -> dict:
    task = latest_sync(db)
    if not task:
        return {"status": "未同步", "last_success_time": None, "last_error": None}
    return {
        "sync_id": task.sync_id,
        "sync_type": task.sync_type,
        "status": task.status,
        "start_time": task.start_time,
        "end_time": task.end_time,
        "last_error": task.error_message,
        "inserted_count": task.inserted_count,
        "updated_count": task.updated_count,
    }


@app.get("/api/accounts")
def accounts(db: Session = Depends(get_db)) -> dict:
    account_rows = {item.account_id: item for item in db.query(Account).order_by(Account.source_name.asc()).all()}
    snapshots = latest_account_snapshots(db)
    snapshot_by_account = {item.account_id: item for item in snapshots}
    items = [_account_to_dict(account_rows.get(item.account_id), item) for item in snapshots]
    for account in account_rows.values():
        if account.account_id not in snapshot_by_account:
            items.append(_account_to_dict(account, None))
    return {
        "items": sorted(items, key=lambda item: (item["source_name"], item["account_id"])),
        "count": len(items),
    }


@app.post("/api/accounts")
def create_account(payload: dict = Body(default_factory=dict), db: Session = Depends(get_db)) -> dict:
    account_id = str(payload.get("account_id") or "").strip() or f"acct_{uuid4().hex[:12]}"
    if db.get(Account, account_id):
        raise HTTPException(status_code=400, detail="Account already exists")
    account = Account(account_id=account_id)
    _apply_account_payload(account, payload)
    account.last_sync_time = datetime.now(timezone.utc)
    db.add(account)
    db.commit()
    return {"account": _account_to_dict(account, None)}


@app.patch("/api/accounts/{account_id}")
def update_account(account_id: str, payload: dict = Body(default_factory=dict), db: Session = Depends(get_db)) -> dict:
    account = db.get(Account, account_id)
    if not account:
        raise HTTPException(status_code=404, detail="Account not found")
    _apply_account_payload(account, payload, allow_account_id=False)
    account.last_sync_time = datetime.now(timezone.utc)
    db.add(account)
    db.commit()
    snapshot = next(iter(latest_account_snapshots(db, account_id)), None)
    return {"account": _account_to_dict(account, snapshot)}


@app.delete("/api/accounts/{account_id}")
def delete_account(account_id: str, confirm_with_data: bool = Query(False), db: Session = Depends(get_db)) -> dict:
    account = db.get(Account, account_id)
    if not account:
        raise HTTPException(status_code=404, detail="Account not found")
    data_counts = _account_data_counts(db, account_id)
    has_data = any(data_counts.values())
    if has_data and not confirm_with_data:
        raise HTTPException(status_code=409, detail={"message": "账户已有数据，删除前需要二次确认", "data_counts": data_counts})
    if has_data:
        db.execute(delete(AccountSnapshot).where(AccountSnapshot.account_id == account_id))
        db.execute(delete(PositionSnapshot).where(PositionSnapshot.account_id == account_id))
        db.execute(delete(Deal).where(Deal.account_id == account_id))
        db.execute(delete(TradeReview).where(TradeReview.account_id == account_id))
    db.delete(account)
    db.commit()
    return {"status": "deleted", "data_counts": data_counts}


@app.post("/api/import/{source}/preview")
async def import_preview(source: str, account_id: str = Query(""), file: UploadFile = File(...)) -> dict:
    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="Empty import file")
    try:
        return preview_import(source, file.filename or "import", content, account_id)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc


@app.post("/api/import/{source}/confirm")
def import_confirm(source: str, payload: dict = Body(default_factory=dict), db: Session = Depends(get_db)) -> dict:
    if payload.get("source_name") and str(payload["source_name"]).replace("-", "_") != source.replace("-", "_"):
        raise HTTPException(status_code=400, detail="Import source mismatch")
    try:
        return confirm_import(db, payload)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc


@app.get("/api/import/excel/template")
def import_excel_template() -> StreamingResponse:
    sheets = [
        (
            "账户资产快照",
            ["account_id", "snapshot_time", "total_assets", "cash", "market_value", "currency", "source_name", "display_name", "institution", "account_type", "today_pnl", "floating_pnl", "raw_note"],
            ["acct_demo", "2026-07-05 15:30:00", 100000, 12000, 88000, "CNY", "manual", "示例账户", "示例券商", "cash", 300, 1200, "示例行可删除"],
        ),
        (
            "持仓快照",
            ["account_id", "snapshot_time", "code", "name", "market", "asset_type", "quantity", "average_cost", "current_price", "market_value", "currency", "profit_loss_ratio", "position_weight", "exchange_rate_to_base", "first_buy_time", "last_trade_time"],
            ["acct_demo", "2026-07-05 15:30:00", "US.AAPL", "Apple", "US", "stock", 10, 180, 200, 2000, "USD", 0.1111, 0.02, 7.2, "2026-01-10 22:30:00", "2026-06-20 22:30:00"],
        ),
        (
            "成交记录",
            ["account_id", "deal_id", "order_id", "code", "side", "price", "quantity", "deal_time", "market", "amount", "currency", "fee", "commission", "tax", "raw_note"],
            ["acct_demo", "deal_demo_001", "order_demo_001", "US.AAPL", "BUY", 180, 10, "2026-01-10 22:30:00", "US", 1800, "USD", 1, 1, 0, "示例行可删除"],
        ),
    ]

    try:
        output = _openpyxl_template(sheets)
    except ImportError:
        output = _xlsx_template_without_openpyxl(sheets)
    output.seek(0)
    headers = {"Content-Disposition": 'attachment; filename="ai-portfolio-import-template.xlsx"'}
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )


def _openpyxl_template(sheets: list[tuple[str, list[str], list[object]]]) -> BytesIO:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    workbook = Workbook()
    workbook.remove(workbook.active)
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for title, headers, sample in sheets:
        sheet = workbook.create_sheet(title)
        sheet.append(headers)
        sheet.append(sample)
        for cell in sheet[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
        for column_index, header in enumerate(headers, start=1):
            width = max(len(str(header)) + 4, 14)
            sheet.column_dimensions[get_column_letter(column_index)].width = min(width, 24)
        sheet.freeze_panes = "A2"

    output = BytesIO()
    workbook.save(output)
    return output


def _xlsx_template_without_openpyxl(sheets: list[tuple[str, list[str], list[object]]]) -> BytesIO:
    output = BytesIO()
    with ZipFile(output, "w", ZIP_DEFLATED) as archive:
        archive.writestr("[Content_Types].xml", _xlsx_content_types(len(sheets)))
        archive.writestr("_rels/.rels", _xlsx_root_relationships())
        archive.writestr("xl/workbook.xml", _xlsx_workbook(sheets))
        archive.writestr("xl/_rels/workbook.xml.rels", _xlsx_workbook_relationships(len(sheets)))
        archive.writestr("xl/styles.xml", _xlsx_styles())
        for index, (_title, headers, sample) in enumerate(sheets, start=1):
            archive.writestr(f"xl/worksheets/sheet{index}.xml", _xlsx_sheet([headers, sample]))
    return output


def _xlsx_content_types(sheet_count: int) -> str:
    sheet_overrides = "".join(
        f'<Override PartName="/xl/worksheets/sheet{index}.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>'
        for index in range(1, sheet_count + 1)
    )
    return (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">'
        '<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>'
        '<Default Extension="xml" ContentType="application/xml"/>'
        '<Override PartName="/xl/workbook.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/>'
        '<Override PartName="/xl/styles.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.styles+xml"/>'
        f"{sheet_overrides}</Types>"
    )


def _xlsx_root_relationships() -> str:
    return (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="xl/workbook.xml"/>'
        "</Relationships>"
    )


def _xlsx_workbook(sheets: list[tuple[str, list[str], list[object]]]) -> str:
    sheet_nodes = "".join(
        f'<sheet name="{escape(title)}" sheetId="{index}" r:id="rId{index}"/>'
        for index, (title, _headers, _sample) in enumerate(sheets, start=1)
    )
    return (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
        'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">'
        f"<sheets>{sheet_nodes}</sheets></workbook>"
    )


def _xlsx_workbook_relationships(sheet_count: int) -> str:
    sheet_relationships = "".join(
        f'<Relationship Id="rId{index}" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" Target="worksheets/sheet{index}.xml"/>'
        for index in range(1, sheet_count + 1)
    )
    style_id = sheet_count + 1
    return (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        f'{sheet_relationships}<Relationship Id="rId{style_id}" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/styles" Target="styles.xml"/>'
        "</Relationships>"
    )


def _xlsx_styles() -> str:
    return (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<styleSheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
        '<fonts count="2"><font><sz val="11"/><name val="Calibri"/></font><font><b/><sz val="11"/><name val="Calibri"/></font></fonts>'
        '<fills count="3"><fill><patternFill patternType="none"/></fill><fill><patternFill patternType="gray125"/></fill><fill><patternFill patternType="solid"><fgColor rgb="FFD9EAF7"/></patternFill></fill></fills>'
        '<borders count="1"><border><left/><right/><top/><bottom/><diagonal/></border></borders>'
        '<cellStyleXfs count="1"><xf numFmtId="0" fontId="0" fillId="0" borderId="0"/></cellStyleXfs>'
        '<cellXfs count="2"><xf numFmtId="0" fontId="0" fillId="0" borderId="0"/><xf numFmtId="0" fontId="1" fillId="2" borderId="0" applyFont="1" applyFill="1"/></cellXfs>'
        "</styleSheet>"
    )


def _xlsx_sheet(rows: list[list[object]]) -> str:
    row_nodes = []
    for row_index, row in enumerate(rows, start=1):
        cell_nodes = []
        for column_index, value in enumerate(row, start=1):
            style = ' s="1"' if row_index == 1 else ""
            cell_nodes.append(_xlsx_cell(column_index, row_index, value, style))
        row_nodes.append(f'<row r="{row_index}">{"".join(cell_nodes)}</row>')
    return (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
        '<sheetViews><sheetView workbookViewId="0"><pane ySplit="1" topLeftCell="A2" activePane="bottomLeft" state="frozen"/></sheetView></sheetViews>'
        f'<sheetData>{"".join(row_nodes)}</sheetData></worksheet>'
    )


def _xlsx_cell(column_index: int, row_index: int, value: object, style: str) -> str:
    reference = f"{_xlsx_column_name(column_index)}{row_index}"
    if isinstance(value, (int, float)):
        return f'<c r="{reference}"{style}><v>{value}</v></c>'
    return f'<c r="{reference}" t="inlineStr"{style}><is><t>{escape(str(value))}</t></is></c>'


def _xlsx_column_name(index: int) -> str:
    name = ""
    while index:
        index, remainder = divmod(index - 1, 26)
        name = chr(65 + remainder) + name
    return name


@app.get("/api/dashboard")
def dashboard(account_id: str = Query("all"), db: Session = Depends(get_db)) -> dict:
    positions = latest_positions(db, account_id)
    position_by_code = {item.code: item for item in positions}
    position_order = {item.code: index for index, item in enumerate(positions)}
    position_codes = set(position_by_code)
    account_snapshots = latest_account_snapshots(db, account_id)
    cards = [card for card in latest_cards(db, limit=max(len(positions), 100)) if card.code in position_codes]
    visible_cards = sorted(cards, key=lambda card: position_order.get(card.code, 10**9))
    base_currency = _portfolio_base_currency(account_snapshots, positions, account_id)
    display_rate_targets = _display_rate_targets(base_currency, account_snapshots, positions)
    display_rates, display_rate_meta = fx_display_rates(db, base_currency, display_rate_targets)
    total_value = _portfolio_market_value(account_snapshots, positions, base_currency, display_rates)
    total_assets = _portfolio_total_assets(account_snapshots, total_value, base_currency, display_rates)
    cash = sum(_money_to_base(item.cash, _snapshot_currency(item), base_currency, display_rates) for item in account_snapshots)
    cash_ratio = cash / total_assets if total_assets else _estimate_cash_ratio(positions)
    position_exposures = _aggregate_positions(positions, total_assets, base_currency, display_rates)
    account_rows = {item.account_id: item for item in db.query(Account).all()}
    return {
        "sync": sync_status(db),
        "portfolio": {
            "account_id": account_id,
            "account_count": len(account_snapshots),
            "position_count": len(positions),
            "total_assets": total_assets,
            "total_position_value": total_value,
            "cash": cash,
            "base_currency": base_currency,
            "display_currencies": list(display_rates.keys()),
            "display_rates": display_rates,
            "display_rate_meta": display_rate_meta,
            "cash_ratio": cash_ratio,
            "max_position_weight": max((item["position_weight"] for item in position_exposures), default=0),
            "max_account_weight": _max_account_weight(account_snapshots, base_currency, display_rates),
            "accounts": [_account_to_dict(account_rows.get(item.account_id), item, base_currency, display_rates) for item in account_snapshots],
        },
        "action_cards": [_card_to_dict(card, position_by_code.get(card.code)) for card in visible_cards],
        "decision_card_state": _decision_card_state(positions, cards),
        "freshness": freshness_summary(db),
        "empty": not positions,
    }


@app.post("/api/decision-cards/generate-ai")
def generate_decision_cards(payload: dict = Body(default_factory=dict), db: Session = Depends(get_db)) -> dict:
    settings = get_settings()
    runtime = active_ai_runtime(db)
    sends_external_context = runtime_can_call_external(runtime)
    if sends_external_context and payload.get("consent_external_ai") is not True:
        raise HTTPException(status_code=400, detail="External AI decision cards require explicit consent")
    scene_config = get_ai_scene_config(db, "decision_cards", str(runtime.get("model") or settings.deepseek_model), DECISION_CARD_SYSTEM_INSTRUCTION, str(runtime.get("provider") or settings.ai_provider))
    model = str(scene_config["model"])
    system_prompt = str(payload.get("system_prompt") or scene_config["system_prompt"])
    return generate_ai_decision_cards(db, model_override=model, system_instruction_override=system_prompt)


@app.get("/api/review")
def review(account_id: str = Query("all"), db: Session = Depends(get_db)) -> dict:
    report = latest_review(db)
    if not report:
        return {"empty": True, "message": "暂无盘后复盘，完成一次同步后生成"}
    return {
        "empty": False,
        "review_id": report.review_id,
        "review_date": report.review_date,
        "portfolio_summary": report.portfolio_summary,
        "advice_summary": report.advice_summary,
        "user_action_summary": report.user_action_summary,
        "result_summary": report.result_summary,
        "next_watchlist": report.next_watchlist,
        "created_at": report.created_at,
    }


@app.get("/api/review/trades")
def trade_reviews(
    code: Optional[str] = None,
    side: Optional[str] = None,
    label: Optional[str] = None,
    account_id: str = Query("all"),
    db: Session = Depends(get_db),
) -> dict:
    return trade_review_payload(db, code=code, side=side, label=label, account_id=account_id)


@app.post("/api/review/trades/refresh-market-data")
def refresh_trade_review_market_data(account_id: str = Query("all"), db: Session = Depends(get_db)) -> dict:
    payload = trade_review_payload(db, account_id=account_id, fetch_market_data=True)
    summary = payload.get("summary", {})
    return {
        "status": "ok",
        "message": "复盘行情已刷新",
        "summary": summary,
    }


@app.patch("/api/review/trades/{review_id}/intent")
def update_trade_review_user_intent(review_id: str, payload: dict = Body(default_factory=dict), db: Session = Depends(get_db)) -> dict:
    review_item = update_trade_review_intent(db, review_id, payload)
    if not review_item:
        raise HTTPException(status_code=404, detail="Trade review not found")
    return _trade_review_update_response(review_item)


def _trade_review_update_response(review_item) -> dict:
    return {
        "review_id": review_item.review_id,
        "user_note": review_item.user_note,
        "intent_tags": review_item.intent_tags,
        "intent_plan": review_item.intent_plan,
        "discipline_label": review_item.discipline_label,
        "ai_commentary": review_item.ai_commentary,
        "updated_at": review_item.updated_at,
    }


@app.get("/api/profile")
def profile(account_id: str = Query("all"), db: Session = Depends(get_db)) -> dict:
    item = latest_profile(db)
    if not item:
        return {"empty": True, "message": "暂无画像，完成一次同步后生成"}
    return {
        "empty": False,
        "profile_id": item.profile_id,
        "generated_at": item.generated_at,
        "confidence": item.confidence,
        "ratios": {
            "核心长期仓": item.core_position_ratio,
            "中期配置仓": item.mid_position_ratio,
            "短期交易仓": item.trade_position_ratio,
            "期权仓": item.option_position_ratio,
        },
        "tags": item.tags,
        "change_reason": item.change_reason,
    }


@app.get("/api/profile/preferences")
def profile_preferences(account_id: str = Query("all"), db: Session = Depends(get_db)) -> dict:
    return preference_to_dict(get_investor_preference(db, account_id), account_id)


@app.patch("/api/profile/preferences")
def save_profile_preferences(
    account_id: str = Query("all"),
    payload: dict = Body(default_factory=dict),
    db: Session = Depends(get_db),
) -> dict:
    preference = upsert_investor_preference(db, account_id, payload)
    return preference_to_dict(preference, account_id)


@app.post("/api/profile/ai-workflows/{workflow_type}")
def create_profile_ai_workflow(
    workflow_type: str,
    account_id: str = Query("all"),
    payload: dict = Body(default_factory=dict),
    db: Session = Depends(get_db),
) -> dict:
    if workflow_type not in WORKFLOW_TYPES:
        raise HTTPException(status_code=400, detail="Unsupported workflow type")
    settings = get_settings()
    runtime = active_ai_runtime(db)
    sends_external_context = runtime_can_call_external(runtime)
    if sends_external_context and payload.get("consent_external_ai") is not True:
        raise HTTPException(status_code=400, detail="External AI workflow requires explicit consent")
    scene_config = get_ai_scene_config(db, scene_key(workflow_type), str(runtime.get("model") or settings.deepseek_model), WORKFLOW_SYSTEM_INSTRUCTION, str(runtime.get("provider") or settings.ai_provider))
    model = str(scene_config["model"])
    planning_config = get_ai_scene_config(db, "planning_agent", str(runtime.get("model") or settings.deepseek_model), "", str(runtime.get("provider") or settings.ai_provider))
    system_prompt = str(payload.get("system_prompt") or scene_config["system_prompt"])
    use_external_model = bool(payload.get("use_external_model", sends_external_context))
    run = create_workflow_run(
        db,
        workflow_type,
        account_id,
        str(payload.get("question") or ""),
        use_external_model=use_external_model,
        model_override=model,
        planning_model_override=str(planning_config["model"]),
        system_instruction_override=system_prompt,
    )
    return {"run": workflow_run_to_dict(run)}


@app.get("/api/profile/ai-workflows")
def profile_ai_workflows(account_id: str = Query("all"), limit: Optional[int] = Query(None, ge=1, le=500), db: Session = Depends(get_db)) -> dict:
    runs = latest_workflow_runs(db, account_id=account_id, limit=limit)
    return {"items": [workflow_run_to_dict(item) for item in runs], "count": len(runs)}


@app.get("/api/profile/ai-workflows/{run_id}")
def profile_ai_workflow_detail(run_id: str, db: Session = Depends(get_db)) -> dict:
    run = db.get(AIWorkflowRun, run_id)
    if not run:
        raise HTTPException(status_code=404, detail="Workflow run not found")
    return {"run": workflow_run_to_dict(run)}


@app.post("/api/profile/ai-workflows/{run_id}/cancel")
def cancel_profile_ai_workflow(run_id: str, db: Session = Depends(get_db)) -> dict:
    try:
        run = cancel_workflow_run(db, run_id)
    except ValueError:
        raise HTTPException(status_code=404, detail="Workflow run not found") from None
    return {"run": workflow_run_to_dict(run)}


@app.delete("/api/profile/ai-workflows/{run_id}")
def delete_profile_ai_workflow(run_id: str, db: Session = Depends(get_db)) -> dict:
    try:
        delete_workflow_run(db, run_id)
    except ValueError:
        raise HTTPException(status_code=404, detail="Workflow run not found") from None
    return {"status": "deleted"}


@app.post("/api/profile/ai-workflows/{run_id}/delete")
def delete_profile_ai_workflow_action(run_id: str, db: Session = Depends(get_db)) -> dict:
    try:
        delete_workflow_run(db, run_id)
    except ValueError:
        raise HTTPException(status_code=404, detail="Workflow run not found") from None
    return {"status": "deleted"}


@app.get("/api/profile/ai-workflows/{run_id}/download")
def download_profile_ai_workflow(run_id: str, db: Session = Depends(get_db)) -> StreamingResponse:
    run = db.get(AIWorkflowRun, run_id)
    if not run:
        raise HTTPException(status_code=404, detail="Workflow run not found")
    markdown = str((run.output or {}).get("markdown") or (run.output or {}).get("partial_markdown") or "")
    if not markdown.strip():
        raise HTTPException(status_code=404, detail="报告正文尚未生成，无法下载")
    filename = workflow_markdown_filename(run)
    return StreamingResponse(
        BytesIO(markdown.encode("utf-8")),
        media_type="text/markdown; charset=utf-8",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{filename}"},
    )


@app.get("/api/profile/ai-workflows/{run_id}/stream")
def profile_ai_workflow_stream(run_id: str) -> StreamingResponse:
    return StreamingResponse(
        stream_workflow_run(run_id),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "X-Accel-Buffering": "no",
            "Connection": "keep-alive",
        },
    )


@app.get("/api/data/status")
def data_status(db: Session = Depends(get_db)) -> dict:
    states = latest_source_states(db)
    return {
        "sync": sync_status(db),
        "freshness": freshness_summary(db),
        "sources": [
            {
                "source_name": item.source_name,
                "account_id": item.account_id,
                "provider": item.provider,
                "data_type": item.data_type,
                "market": item.market,
                "status": item.status,
                "last_success_time": item.last_success_time,
                "last_error": item.last_error,
                "freshness_seconds": item.freshness_seconds,
                "updated_at": item.updated_at,
            }
            for item in states
        ],
    }


@app.get("/api/data/accounts/{account_id}/overview")
def account_data_overview(account_id: str, db: Session = Depends(get_db)) -> dict:
    account = db.get(Account, account_id)
    if not account:
        raise HTTPException(status_code=404, detail="Account not found")
    snapshot = next(iter(latest_account_snapshots(db, account_id)), None)
    positions = latest_positions(db, account_id)
    position_codes = [item.code for item in positions if item.code and not item.missing_market_code]
    markets = _account_data_markets(account, positions)
    latest_quote_time = None
    latest_news_time = None
    if position_codes:
        latest_quote_time = db.scalar(select(func.max(QuoteSummary.quote_time)).where(QuoteSummary.code.in_(position_codes)))
        latest_news_time = db.scalar(select(func.max(NewsItem.fetched_at)).where(NewsItem.code.in_(position_codes)))
    deals = list(
        db.scalars(
            select(Deal)
            .where(Deal.account_id == account_id)
            .order_by(Deal.deal_time.desc().nullslast(), Deal.id.desc())
            .limit(200)
        ).all()
    )
    return {
        "account": _account_to_dict(account, snapshot),
        "asset_snapshot": _asset_snapshot_to_dict(snapshot) if snapshot else None,
        "positions": [_position_to_dict(item) for item in positions],
        "deals": [_deal_to_dict(item) for item in deals],
        "updated_at": {
            "account": snapshot.snapshot_time if snapshot else account.last_sync_time,
            "position": max((item.snapshot_time for item in positions), default=None),
            "deal": max((item.deal_time for item in deals if item.deal_time), default=None),
            "quote": latest_quote_time,
            "news": latest_news_time,
        },
        "provider_states": _account_provider_states(db, account, markets),
    }


@app.post("/api/data/accounts/{account_id}/providers/check")
def check_account_data_provider(account_id: str, payload: dict = Body(default_factory=dict), db: Session = Depends(get_db)) -> dict:
    account = db.get(Account, account_id)
    if not account:
        raise HTTPException(status_code=404, detail="Account not found")
    data_type = str(payload.get("data_type") or "quote").strip().lower()
    raw_market = str(payload.get("market") or "").strip()
    positions = latest_positions(db, account_id)
    markets = _account_data_markets(account, positions)
    market = normalize_market(raw_market) if raw_market else (markets[0] if markets else "US")
    registry = provider_registry()
    provider_name = str(payload.get("provider") or "").strip().lower()
    preferred_provider = _account_data_provider_preference(account, data_type)
    requested_provider = provider_name or preferred_provider
    provider = registry.get(requested_provider) if requested_provider else None
    if provider and not _provider_supports(provider, data_type, market):
        status = "unsupported"
        message = "当前 provider 不支持该市场或数据类型"
        state_key = f"{account_id}:{provider.name}:{data_type}:{market}"
        mark_source_state(db, state_key, status, message, 0, account_id=account_id, provider=provider.name, data_type=data_type, market=market)
        db.commit()
        return {
            "provider": provider.name,
            "provider_label": provider.label,
            "data_type": data_type,
            "market": market,
            "status": status,
            "message": message,
            "checked_at": datetime.now(timezone.utc),
            "license_note": provider.license_note,
            "provider_states": _account_provider_states(db, account, markets),
        }
    provider = provider or registry.choose(data_type=data_type, market=market, broker_provider=_account_broker_provider(account))
    if not provider:
        raise HTTPException(status_code=400, detail="Unsupported provider for this market and data type")
    health = provider.health()
    status = "available" if health.status in {"available", "configured"} else health.status
    state_key = f"{account_id}:{provider.name}:{data_type}:{market}"
    mark_source_state(
        db,
        state_key,
        status,
        health.message,
        0,
        account_id=account_id,
        provider=provider.name,
        data_type=data_type,
        market=market,
    )
    db.commit()
    return {
        "provider": provider.name,
        "provider_label": provider.label,
        "data_type": data_type,
        "market": market,
        "status": status,
        "message": health.message,
        "checked_at": health.checked_at,
        "license_note": provider.license_note,
        "provider_states": _account_provider_states(db, account, markets),
    }


@app.post("/api/data/accounts/{account_id}/market-data/pull")
def pull_account_data_source(account_id: str, payload: dict = Body(default_factory=dict), db: Session = Depends(get_db)) -> dict:
    account = db.get(Account, account_id)
    if not account:
        raise HTTPException(status_code=404, detail="Account not found")
    data_type = str(payload.get("data_type") or "quote").strip().lower()
    try:
        result = pull_account_market_data(db, account, data_type)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    except Exception as exc:
        raise HTTPException(status_code=502, detail=str(exc)[:500]) from exc
    positions = latest_positions(db, account_id)
    return {
        **result,
        "provider_states": _account_provider_states(db, account, _account_data_markets(account, positions)),
    }


@app.post("/api/data/accounts/{account_id}/positions")
def save_account_position(account_id: str, request: PositionSnapshotSaveRequest, db: Session = Depends(get_db)) -> dict:
    account = db.get(Account, account_id)
    if not account:
        raise HTTPException(status_code=404, detail="Account not found")
    code = request.code.strip()
    if not code:
        raise HTTPException(status_code=400, detail="Position code is required")
    if request.quantity < 0:
        raise HTTPException(status_code=400, detail="Quantity cannot be negative")

    snapshot_time = _as_utc_naive(request.snapshot_time)
    original_time = _as_utc_naive(request.original_snapshot_time) if request.original_snapshot_time else snapshot_time
    original_code = (request.original_code or code).strip()
    position = _find_position_for_manual_save(db, account_id, original_code, original_time, code, snapshot_time)

    currency = (request.currency or request.normalized_currency or account.base_currency or "CNY").strip().upper()
    normalized_currency = (request.normalized_currency or currency).strip().upper()
    raw_market_value = request.market_value
    if raw_market_value is None:
        raw_market_value = request.quantity * request.current_price
    normalized_market_value = request.normalized_market_value
    if normalized_market_value is None:
        rate = request.exchange_rate_to_base or 1
        normalized_market_value = raw_market_value * rate
    profit_loss_ratio = request.profit_loss_ratio
    if profit_loss_ratio is None:
        profit_loss_ratio = ((request.current_price - request.average_cost) / request.average_cost) if request.average_cost else 0

    if position is None:
        position = PositionSnapshot(account_id=account_id)
    position.code = code
    position.name = request.name.strip() or code
    raw_market = request.market.strip() or (code.split(".", 1)[0] if "." in code else "")
    position.market = normalize_market(raw_market)
    position.asset_type = request.asset_type.strip() or "stock"
    position.quantity = request.quantity
    position.average_cost = request.average_cost
    position.current_price = request.current_price
    position.raw_market_value = raw_market_value
    position.raw_currency = currency
    position.normalized_market_value = normalized_market_value
    position.normalized_currency = normalized_currency
    position.exchange_rate_to_base = request.exchange_rate_to_base
    position.position_weight = request.position_weight if request.position_weight is not None else 0
    position.profit_loss_ratio = profit_loss_ratio
    position.position_layer = request.position_layer or "中期配置仓"
    position.layer_source = "user"
    position.layer_confidence = "高"
    position.layer_reason = "用户手动维护持仓快照"
    position.missing_market_code = not bool(raw_market)
    position.snapshot_time = snapshot_time
    position.sync_id = f"manual_position_{account_id}"
    db.add(position)
    db.flush()
    _recalculate_account_position_weights(db, account_id)
    db.commit()
    return {"status": "saved", "overview": account_data_overview(account_id, db)}


@app.delete("/api/data/accounts/{account_id}/positions")
def delete_account_position(account_id: str, code: str = Query(...), snapshot_time: datetime = Query(...), db: Session = Depends(get_db)) -> dict:
    account = db.get(Account, account_id)
    if not account:
        raise HTTPException(status_code=404, detail="Account not found")
    target_time = _as_utc_naive(snapshot_time)
    position = db.scalar(
        select(PositionSnapshot).where(
            PositionSnapshot.account_id == account_id,
            PositionSnapshot.code == code,
            PositionSnapshot.snapshot_time == target_time,
        )
    )
    if not position:
        raise HTTPException(status_code=404, detail="Position not found")
    db.delete(position)
    db.flush()
    _recalculate_account_position_weights(db, account_id)
    db.commit()
    return {"status": "deleted", "overview": account_data_overview(account_id, db)}


@app.post("/api/data/accounts/{account_id}/deals")
def save_account_deal(account_id: str, request: DealSaveRequest, db: Session = Depends(get_db)) -> dict:
    account = db.get(Account, account_id)
    if not account:
        raise HTTPException(status_code=404, detail="Account not found")
    deal_id = request.deal_id.strip()
    if not deal_id:
        raise HTTPException(status_code=400, detail="Deal ID is required")
    code = request.code.strip()
    if not code:
        raise HTTPException(status_code=400, detail="Deal code is required")
    if request.price < 0 or request.quantity < 0:
        raise HTTPException(status_code=400, detail="Price and quantity cannot be negative")

    original_id = (request.original_deal_id or deal_id).strip()
    original = db.scalar(select(Deal).where(Deal.account_id == account_id, Deal.deal_id == original_id))
    target = db.scalar(select(Deal).where(Deal.account_id == account_id, Deal.deal_id == deal_id))
    if original and target and original.id != target.id:
        db.delete(original)
        db.flush()
        deal = target
    else:
        deal = original or target or Deal(account_id=account_id)

    deal.deal_id = deal_id
    deal.order_id = request.order_id.strip()
    deal.code = code
    deal.side = request.side.strip()
    deal.price = request.price
    deal.quantity = request.quantity
    deal.deal_time = _as_utc_naive(request.deal_time) if request.deal_time else None
    deal.market = request.market.strip().upper()
    deal.account_id = account_id
    deal.raw_payload = {
        "source": "manual",
        "amount": request.price * request.quantity,
    }
    db.add(deal)
    db.commit()
    return {"status": "saved", "overview": account_data_overview(account_id, db)}


@app.delete("/api/data/accounts/{account_id}/deals/{deal_id}")
def delete_account_deal(account_id: str, deal_id: str, db: Session = Depends(get_db)) -> dict:
    account = db.get(Account, account_id)
    if not account:
        raise HTTPException(status_code=404, detail="Account not found")
    deal = db.scalar(select(Deal).where(Deal.account_id == account_id, Deal.deal_id == deal_id))
    if not deal:
        raise HTTPException(status_code=404, detail="Deal not found")
    db.delete(deal)
    db.commit()
    return {"status": "deleted", "overview": account_data_overview(account_id, db)}


@app.get("/api/positions")
def positions(layer: Optional[str] = None, account_id: str = Query("all"), db: Session = Depends(get_db)) -> dict:
    items = latest_positions(db, account_id)
    if layer:
        items = [item for item in items if item.position_layer == layer]
    if account_id == "all":
        account_snapshots = latest_account_snapshots(db, account_id)
        base_currency = _portfolio_base_currency(account_snapshots, items, account_id)
        display_rates, _meta = fx_display_rates(db, base_currency, _display_rate_targets(base_currency, account_snapshots, items))
        total_value = _portfolio_market_value(account_snapshots, items, base_currency, display_rates)
        total_assets = _portfolio_total_assets(account_snapshots, total_value, base_currency, display_rates)
        payload = _aggregate_positions(items, total_assets, base_currency, display_rates)
    else:
        payload = [_position_to_dict(item) for item in items]
    return {"items": payload, "count": len(payload)}


@app.get("/api/positions/kline/{code}")
def position_kline(
    code: str,
    account_id: str = Query("all"),
    ktype: str = Query("K_DAY"),
    count: int = Query(90, ge=5, le=300),
    db: Session = Depends(get_db),
) -> dict:
    items = [item for item in latest_positions(db, account_id) if item.code == code]
    if not items:
        raise HTTPException(status_code=404, detail="Position not found")
    if items[0].missing_market_code:
        return {"code": code, "status": "missing", "message": "该持仓缺少真实证券代码", "items": []}
    allowed_ktype = ktype if ktype in {"K_DAY", "K_WEEK", "K_MON"} else "K_DAY"
    snapshot_time = db.scalar(
        select(KlineSnapshot.snapshot_time)
        .where(KlineSnapshot.code == code, KlineSnapshot.period == allowed_ktype)
        .order_by(KlineSnapshot.snapshot_time.desc())
        .limit(1)
    )
    if not snapshot_time:
        return {"code": code, "ktype": allowed_ktype, "status": "missing", "message": "未同步 K 线，技术分析已降级。", "items": []}
    rows = list(
        db.scalars(
            select(KlineSnapshot)
            .where(KlineSnapshot.code == code, KlineSnapshot.period == allowed_ktype, KlineSnapshot.snapshot_time == snapshot_time)
            .order_by(KlineSnapshot.time_key.desc())
            .limit(count)
        ).all()
    )
    return {
        "code": code,
        "ktype": allowed_ktype,
        "status": "available" if rows else "missing",
        "message": "" if rows else "未同步 K 线，技术分析已降级。",
        "snapshot_time": snapshot_time.isoformat(),
        "items": [
            {"time_key": row.time_key, "open": row.open, "close": row.close, "high": row.high, "low": row.low, "volume": row.volume, "turnover": row.turnover}
            for row in reversed(rows)
        ],
    }


@app.get("/api/positions/{code:path}")
def position_detail(code: str, account_id: str = Query("all"), db: Session = Depends(get_db)) -> dict:
    items = [item for item in latest_positions(db, account_id) if item.code == code]
    if not items:
        raise HTTPException(status_code=404, detail="Position not found")
    if account_id == "all":
        account_snapshots = latest_account_snapshots(db, account_id)
        base_currency = _portfolio_base_currency(account_snapshots, items, account_id)
        display_rates, _meta = fx_display_rates(db, base_currency, _display_rate_targets(base_currency, account_snapshots, items))
        total_value = _portfolio_market_value(account_snapshots, items, base_currency, display_rates)
        total_assets = _portfolio_total_assets(account_snapshots, total_value, base_currency, display_rates)
        position = _aggregate_positions(items, total_assets, base_currency, display_rates)[0]
    else:
        position = items[0]
    cards = [card for card in latest_cards(db, limit=100, latest_per_code=False) if card.code == code]
    cards = _merge_duplicate_detail_cards(cards)
    ai_analysis = latest_ai_analysis(db, code)
    news = recent_news_for_code(db, code, days=3, limit=12)
    _record_user_action(db, code, "查看标的详情", cards[0].card_id if cards else None)
    return {
        "position": _position_to_dict(position),
        "account_positions": [_position_to_dict(item) for item in items],
        "cards": [_card_to_dict(card, position) for card in cards],
        "news": [_news_to_dict(item) for item in news],
        "ai_analysis": _ai_analysis_to_dict(ai_analysis) if ai_analysis else None,
    }


@app.post("/api/positions/{code:path}/ai-analysis")
def create_position_ai_analysis(
    code: str,
    account_id: str = Query("all"),
    payload: dict = Body(default_factory=dict),
    db: Session = Depends(get_db),
) -> dict:
    items = [item for item in latest_positions(db, account_id) if item.code == code]
    if not items:
        raise HTTPException(status_code=404, detail="Position not found")
    settings = get_settings()
    runtime = active_ai_runtime(db)
    sends_external_context = runtime_can_call_external(runtime)
    if sends_external_context and payload.get("consent_external_ai") is not True:
        raise HTTPException(status_code=400, detail="External AI analysis requires explicit consent")
    position = items[0]
    scene_config = get_ai_scene_config(db, "ai_analysis", str(runtime.get("model") or settings.deepseek_model), AI_ANALYSIS_SYSTEM_INSTRUCTION, str(runtime.get("provider") or settings.ai_provider))
    model = str(scene_config["model"])
    system_prompt = str(payload.get("system_prompt") or scene_config["system_prompt"])
    analysis = generate_ai_analysis(db, position, model_override=model, system_instruction_override=system_prompt)
    _record_user_action(db, code, "生成AI分析", None)
    return {"ai_analysis": _ai_analysis_to_dict(analysis)}


@app.patch("/api/positions/{code:path}/layer")
def update_layer(code: str, request: LayerOverrideRequest, db: Session = Depends(get_db)) -> dict:
    allowed = {"核心长期仓", "中期配置仓", "短期交易仓", "期权仓", "遗留观察仓"}
    if request.position_layer not in allowed:
        raise HTTPException(status_code=400, detail="Unsupported position layer")
    override = db.get(PositionLayerOverride, code) or PositionLayerOverride(code=code)
    override.position_layer = request.position_layer
    override.reason = request.reason
    from datetime import datetime, timezone

    override.updated_at = datetime.now(timezone.utc)
    db.add(override)
    latest_items = [item for item in latest_positions(db) if item.code == code]
    for item in latest_items:
        item.position_layer = request.position_layer
        item.layer_source = "user"
        item.layer_confidence = "高"
        item.layer_reason = request.reason or "用户手动修正优先"
        db.add(item)
    db.commit()
    return {
        "code": code,
        "position_layer": request.position_layer,
        "updated_positions": len(latest_items),
        "message": "仓位类型已保存，并会在后续同步时继续优先使用用户修正",
    }


@app.post("/api/data/delete-local")
def delete_data(request: DeleteLocalDataRequest, db: Session = Depends(get_db)) -> dict:
    if request.confirmation != "DELETE_LOCAL_DATA":
        raise HTTPException(status_code=400, detail="Confirmation mismatch")
    delete_local_data(db)
    return {"status": "deleted"}


def _apply_account_payload(account: Account, payload: dict, allow_account_id: bool = True) -> None:
    if "import_modes" in payload:
        normalized_modes = _normalize_import_modes(payload.get("import_modes"))
        if normalized_modes:
            payload = {**payload, "import_mode": ",".join(normalized_modes)}
    elif "import_mode" in payload:
        normalized_modes = _normalize_import_modes(payload.get("import_mode"))
        if normalized_modes:
            payload = {**payload, "import_mode": ",".join(normalized_modes)}
    if "position_import_modes" in payload:
        normalized_modes = _normalize_import_modes(payload.get("position_import_modes"))
        if normalized_modes:
            payload = {**payload, "position_import_modes": ",".join(normalized_modes)}
    if "review_import_modes" in payload:
        normalized_modes = _normalize_import_modes(payload.get("review_import_modes"))
        if normalized_modes:
            payload = {**payload, "review_import_modes": ",".join(normalized_modes)}
    allowed = {
        "account_id",
        "source_name",
        "broker_provider",
        "display_name",
        "institution",
        "import_mode",
        "position_import_modes",
        "review_import_modes",
        "market_data_provider",
        "news_data_provider",
        "account_type",
        "base_currency",
        "markets",
        "enabled",
    }
    for key in allowed:
        if key == "account_id" and not allow_account_id:
            continue
        if key not in payload:
            continue
        value = payload[key]
        if key == "markets":
            if isinstance(value, str):
                value = [item.strip().upper() for item in value.split(",") if item.strip()]
            elif not isinstance(value, list):
                value = []
        if key == "enabled":
            value = bool(value)
        if key in {"market_data_provider", "news_data_provider"}:
            value = str(value or "").strip().lower()
        setattr(account, key, value)
    if not account.source_name:
        account.source_name = "manual"
    if not account.import_mode:
        account.import_mode = "manual"
    legacy_modes = ",".join(_account_import_modes(account.import_mode, _account_broker_provider(account)))
    if not getattr(account, "position_import_modes", ""):
        account.position_import_modes = legacy_modes
    if not getattr(account, "review_import_modes", ""):
        account.review_import_modes = legacy_modes
    if not account.base_currency:
        account.base_currency = "CNY"


def _account_data_counts(db: Session, account_id: str) -> dict:
    return {
        "asset_snapshots": db.scalar(select(func.count()).select_from(AccountSnapshot).where(AccountSnapshot.account_id == account_id)) or 0,
        "positions": db.scalar(select(func.count()).select_from(PositionSnapshot).where(PositionSnapshot.account_id == account_id)) or 0,
        "deals": db.scalar(select(func.count()).select_from(Deal).where(Deal.account_id == account_id)) or 0,
    }


def _account_data_markets(account: Account, positions: list[PositionSnapshot]) -> list[str]:
    supported_markets = {"US", "HK", "CN"}
    markets = set()
    for item in positions:
        if item.market:
            market = normalize_market(item.market)
            if market in supported_markets:
                markets.add(market)
            continue
        code = str(item.code or "").strip()
        if "." in code:
            market = normalize_market(code.split(".", 1)[0])
            if market in supported_markets:
                markets.add(market)
    raw_account_markets = account.markets or []
    if isinstance(raw_account_markets, str):
        raw_account_markets = [item.strip() for item in raw_account_markets.split(",") if item.strip()]
    markets.update(market for market in (normalize_market(item) for item in raw_account_markets if item) if market in supported_markets)
    if not markets:
        base_currency = str(account.base_currency or "").upper()
        broker_provider = _account_broker_provider(account)
        if base_currency.startswith("HK"):
            markets.add("HK")
        elif base_currency in {"CNY", "CNH", "RMB"}:
            markets.add("CN")
        elif broker_provider == "futu":
            markets.add("HK")
        else:
            markets.add("US")
    return sorted(item for item in markets if item)


def _account_provider_states(db: Session, account: Account, markets: list[str]) -> list[dict]:
    stored_rows = {
        (item.data_type, item.market, item.provider): item
        for item in db.scalars(
            select(DataSourceState).where(
                DataSourceState.account_id == account.account_id,
                DataSourceState.data_type.in_(["quote", "news", "announcement", "filing"]),
            )
        ).all()
    }
    rows = []
    for item in account_provider_statuses(
        _account_broker_provider(account),
        markets,
        getattr(account, "market_data_provider", ""),
        getattr(account, "news_data_provider", ""),
    ):
        stored = stored_rows.get((item["data_type"], item["market"], item["provider"]))
        if stored:
            rows.append(
                {
                    **item,
                    "status": stored.status,
                    "message": stored.last_error or item["message"],
                    "last_success_time": stored.last_success_time,
                    "freshness_seconds": stored.freshness_seconds,
                    "checked_at": stored.updated_at,
                }
            )
        else:
            rows.append({**item, "last_success_time": None, "freshness_seconds": 0})
    return rows


def _asset_snapshot_to_dict(item) -> dict:
    return {
        "account_id": item.account_id,
        "total_assets": item.total_assets,
        "cash": item.cash,
        "market_value": item.market_value,
        "currency": _snapshot_currency(item),
        "raw_currency_values": item.raw_currency_values,
        "snapshot_time": item.snapshot_time,
        "sync_id": item.sync_id,
    }


def _deal_to_dict(item: Deal) -> dict:
    return {
        "deal_id": item.deal_id,
        "order_id": item.order_id,
        "code": item.code,
        "side": item.side,
        "price": item.price,
        "quantity": item.quantity,
        "deal_time": item.deal_time,
        "market": item.market,
        "account_id": item.account_id,
        "raw_payload": item.raw_payload,
    }


def _as_utc_naive(value: datetime) -> datetime:
    if value.tzinfo:
        return value.astimezone(timezone.utc).replace(tzinfo=None)
    return value


def _find_position_for_manual_save(
    db: Session,
    account_id: str,
    original_code: str,
    original_time: datetime,
    code: str,
    snapshot_time: datetime,
) -> PositionSnapshot | None:
    original = db.scalar(
        select(PositionSnapshot).where(
            PositionSnapshot.account_id == account_id,
            PositionSnapshot.code == original_code,
            PositionSnapshot.snapshot_time == original_time,
        )
    )
    target = db.scalar(
        select(PositionSnapshot).where(
            PositionSnapshot.account_id == account_id,
            PositionSnapshot.code == code,
            PositionSnapshot.snapshot_time == snapshot_time,
        )
    )
    if original and target and original.id != target.id:
        db.delete(original)
        db.flush()
        return target
    if original or target:
        return original or target

    candidates = list(
        db.scalars(
            select(PositionSnapshot).where(
                PositionSnapshot.account_id == account_id,
                PositionSnapshot.code == original_code,
            )
        ).all()
    )
    if not candidates:
        return None
    nearest = min(candidates, key=lambda item: abs(item.snapshot_time - original_time))
    if abs(nearest.snapshot_time - original_time) <= timedelta(hours=12):
        return nearest
    return None


def _recalculate_account_position_weights(db: Session, account_id: str) -> None:
    snapshot = next(iter(latest_account_snapshots(db, account_id)), None)
    latest = latest_positions(db, account_id)
    total_assets = snapshot.total_assets if snapshot else sum(item.normalized_market_value for item in latest)
    if total_assets <= 0:
        total_assets = sum(item.normalized_market_value for item in latest) or 1
    for item in latest:
        item.position_weight = item.normalized_market_value / total_assets
        db.add(item)


def _position_to_dict(item) -> dict:
    if isinstance(item, dict):
        return item
    return {
        "account_id": item.account_id,
        "code": item.code,
        "name": item.name,
        "market": item.market,
        "asset_type": item.asset_type,
        "quantity": item.quantity,
        "current_price": item.current_price,
        "average_cost": item.average_cost,
        "raw_market_value": item.raw_market_value,
        "raw_currency": item.raw_currency,
        "normalized_market_value": item.normalized_market_value,
        "normalized_currency": item.normalized_currency,
        "exchange_rate_to_base": item.exchange_rate_to_base,
        "position_weight": item.position_weight,
        "profit_loss_ratio": item.profit_loss_ratio,
        "position_layer": item.position_layer,
        "layer_source": item.layer_source,
        "layer_confidence": item.layer_confidence,
        "layer_reason": item.layer_reason,
        "missing_market_code": item.missing_market_code,
        "snapshot_time": item.snapshot_time,
    }


def _aggregate_positions(
    items,
    total_assets: float = 0,
    base_currency: str = "",
    display_rates: dict[str, float] | None = None,
) -> list[dict]:
    groups: dict[str, list] = {}
    for item in items:
        groups.setdefault(item.code, []).append(item)
    aggregated = []
    for code, group in groups.items():
        first = group[0]
        total_value = sum(_position_market_value_base(item, base_currency, display_rates) for item in group)
        total_quantity = sum(item.quantity for item in group)
        total_cost = sum(item.average_cost * item.quantity for item in group)
        cost_basis = total_cost / total_quantity if total_quantity else first.average_cost
        account_positions = [
            {
                "account_id": item.account_id,
                "market_value": _position_market_value_base(item, base_currency, display_rates),
                "currency": base_currency or item.normalized_currency,
                "quantity": item.quantity,
                "weight": item.position_weight,
            }
            for item in group
        ]
        portfolio_weight = total_value / total_assets if total_assets else sum(item.position_weight for item in group)
        aggregated.append(
            {
                "account_id": "all",
                "code": code,
                "name": first.name,
                "market": first.market,
                "asset_type": first.asset_type,
                "quantity": total_quantity,
                "current_price": first.current_price,
                "average_cost": cost_basis,
                "raw_market_value": total_value,
                "raw_currency": base_currency or first.normalized_currency,
                "normalized_market_value": total_value,
                "normalized_currency": base_currency or first.normalized_currency,
                "exchange_rate_to_base": first.exchange_rate_to_base,
                "position_weight": portfolio_weight,
                "profit_loss_ratio": _weighted_profit_loss(group, base_currency, display_rates),
                "position_layer": first.position_layer,
                "layer_source": first.layer_source,
                "layer_confidence": first.layer_confidence,
                "layer_reason": first.layer_reason,
                "missing_market_code": any(item.missing_market_code for item in group),
                "snapshot_time": max(item.snapshot_time for item in group),
                "account_positions": account_positions,
                "account_count": len(group),
            }
        )
    return sorted(aggregated, key=lambda item: item["normalized_market_value"], reverse=True)


def _position_market_value_base(item, base_currency: str = "", display_rates: dict[str, float] | None = None) -> float:
    return _money_to_base(item.normalized_market_value, item.normalized_currency, base_currency or item.normalized_currency, display_rates)


def _weighted_profit_loss(items, base_currency: str = "", display_rates: dict[str, float] | None = None) -> float:
    total = sum(abs(_position_market_value_base(item, base_currency, display_rates)) for item in items)
    if not total:
        return 0
    return (
        sum(item.profit_loss_ratio * abs(_position_market_value_base(item, base_currency, display_rates)) for item in items)
        / total
    )


def _card_to_dict(card: DecisionCard, position=None) -> dict:
    position_snapshot_time = _position_snapshot_time(position)
    needs_regeneration = _is_position_newer_than_card(position_snapshot_time, card.data_time)
    return {
        "card_id": card.card_id,
        "code": card.code,
        "position_layer": card.position_layer,
        "recommendation": card.recommendation,
        "confidence": card.confidence,
        "reasons": card.reasons,
        "risks": card.risks,
        "key_prices": card.key_prices,
        "status": card.status,
        "priority": card.priority,
        "data_time": card.data_time,
        "action_required": card.action_required,
        "generation_source": card.generation_source,
        "model": card.model,
        "generated_at": card.generated_at or card.created_at,
        "input_version": card.input_version or card.data_version,
        "analysis_framework": card.analysis_framework,
        "missing_data": card.missing_data,
        "invalid_conditions": card.invalid_conditions,
        "needs_regeneration": needs_regeneration,
        "merged_count": getattr(card, "merged_count", 1),
        "merged_first_data_time": getattr(card, "merged_first_data_time", card.data_time),
        "merged_last_data_time": getattr(card, "merged_last_data_time", card.data_time),
    }


def _is_position_newer_than_card(position_snapshot_time: datetime | None, card_data_time: datetime | None) -> bool:
    if not position_snapshot_time or not card_data_time:
        return False
    if position_snapshot_time.tzinfo and not card_data_time.tzinfo:
        position_snapshot_time = position_snapshot_time.replace(tzinfo=None)
    if card_data_time.tzinfo and not position_snapshot_time.tzinfo:
        card_data_time = card_data_time.replace(tzinfo=None)
    return position_snapshot_time > card_data_time


def _position_snapshot_time(position) -> datetime | None:
    if not position:
        return None
    if isinstance(position, dict):
        value = position.get("snapshot_time")
    else:
        value = getattr(position, "snapshot_time", None)
    if isinstance(value, datetime):
        return value
    if isinstance(value, str):
        try:
            return datetime.fromisoformat(value)
        except ValueError:
            return None
    return None


def _news_to_dict(item: NewsItem) -> dict:
    return {
        "news_id": item.news_id,
        "code": item.code,
        "provider": item.provider,
        "market": item.market,
        "news_type": item.news_type,
        "title": item.title,
        "news_sub_type": item.news_sub_type,
        "source": item.source,
        "publish_time": item.publish_time,
        "view_count": item.view_count,
        "related_securities": item.related_securities,
        "url": item.url,
        "fetched_at": item.fetched_at,
    }


def _merge_duplicate_detail_cards(cards: list[DecisionCard]) -> list[DecisionCard]:
    groups: dict[str, list[DecisionCard]] = {}
    for card in cards:
        groups.setdefault(_detail_card_merge_key(card), []).append(card)

    merged = []
    for group in groups.values():
        sorted_group = sorted(group, key=lambda item: item.created_at, reverse=True)
        latest = sorted_group[0]
        data_times = [item.data_time for item in sorted_group]
        latest.merged_count = len(sorted_group)
        latest.merged_first_data_time = min(data_times)
        latest.merged_last_data_time = max(data_times)
        merged.append(latest)
    return sorted(merged, key=lambda item: item.created_at, reverse=True)


def _detail_card_merge_key(card: DecisionCard) -> str:
    payload = {
        "code": card.code,
        "position_layer": card.position_layer,
        "recommendation": card.recommendation,
        "priority": card.priority,
        "confidence": card.confidence,
        "status": card.status,
        "generation_source": card.generation_source,
        "model": card.model,
        "reasons": card.reasons,
        "risks": card.risks,
        "invalid_conditions": card.invalid_conditions,
        "missing_data": card.missing_data,
    }
    return json.dumps(payload, ensure_ascii=False, sort_keys=True)


def _alert_to_dict(alert: Alert) -> dict:
    return {
        "alert_id": alert.alert_id,
        "code": alert.code,
        "alert_type": alert.alert_type,
        "trigger_condition": alert.trigger_condition,
        "priority": alert.priority,
        "message": alert.message,
        "created_at": alert.created_at,
        "read_status": alert.read_status,
        "related_card_id": alert.related_card_id,
    }


def _ai_analysis_to_dict(analysis) -> dict:
    return {
        "analysis_id": analysis.analysis_id,
        "code": analysis.code,
        "provider": analysis.provider,
        "model": analysis.model,
        "output": normalize_ai_output(analysis.output),
        "status": analysis.status,
        "error_message": analysis.error_message,
        "data_version": analysis.data_version,
        "created_at": analysis.created_at,
    }


def _record_user_action(db: Session, code: str, action_type: str, related_card_id: Optional[str]) -> UserAction:
    from datetime import datetime, timezone
    from uuid import uuid4

    action = UserAction(
        action_id=f"action_{uuid4().hex}",
        code=code,
        action_type=action_type,
        related_card_id=related_card_id,
        note="",
        created_at=datetime.now(timezone.utc),
    )
    db.add(action)
    db.commit()
    return action


def _estimate_cash_ratio(positions) -> float:
    if not positions:
        return 0
    exposure = sum(item.position_weight for item in positions)
    return max(0, 1 - exposure)


def _base_currency(positions) -> str:
    for item in positions:
        if item.normalized_currency:
            return item.normalized_currency
    return "HKD"


def _account_market_value(account_snapshots) -> float:
    return sum(item.market_value for item in account_snapshots)


def _account_total_assets(account_snapshots) -> float:
    return sum(item.total_assets for item in account_snapshots)


def _account_cash_ratio(account_snapshots) -> float:
    total_assets = sum(item.total_assets for item in account_snapshots)
    cash = sum(item.cash for item in account_snapshots)
    return cash / total_assets if total_assets else 0


def _account_base_currency(account_snapshots) -> str:
    for snapshot in account_snapshots:
        currency = (snapshot.raw_currency_values or {}).get("currency")
        if currency:
            return str(currency)
    return ""


def _portfolio_base_currency(account_snapshots, positions, account_id: str) -> str:
    if account_id != "all":
        return _account_base_currency(account_snapshots) or _base_currency(positions)
    return "CNY"


def _portfolio_market_value(account_snapshots, positions, base_currency: str, display_rates: dict[str, float]) -> float:
    if account_snapshots:
        return sum(_money_to_base(item.market_value, _snapshot_currency(item), base_currency, display_rates) for item in account_snapshots)
    return sum(_money_to_base(item.normalized_market_value, item.normalized_currency, base_currency, display_rates) for item in positions)


def _portfolio_total_assets(account_snapshots, fallback: float, base_currency: str, display_rates: dict[str, float]) -> float:
    if account_snapshots:
        return sum(_money_to_base(item.total_assets, _snapshot_currency(item), base_currency, display_rates) for item in account_snapshots)
    return fallback


def _snapshot_currency(snapshot) -> str:
    return str((snapshot.raw_currency_values or {}).get("currency") or "")


def _money_to_base(value: float, currency: str, base_currency: str, display_rates: dict[str, float] | None = None) -> float:
    currency = (currency or base_currency or "CNY").upper()
    base_currency = (base_currency or currency).upper()
    if currency == base_currency:
        return value
    rate = (display_rates or {}).get(currency)
    if rate:
        return value / rate
    return value * _currency_rate_to_cny(currency) / _currency_rate_to_cny(base_currency)


def _currency_rate_to_cny(currency: str) -> float:
    rates = {"CNY": 1.0, "CNH": 1.0, "USD": 7.2, "HKD": 0.92}
    return rates.get(currency.upper(), 1.0)


def _max_account_weight(account_snapshots, base_currency: str, display_rates: dict[str, float]) -> float:
    total = sum(_money_to_base(item.total_assets, _snapshot_currency(item), base_currency, display_rates) for item in account_snapshots)
    if not total:
        return 0
    return max((_money_to_base(item.total_assets, _snapshot_currency(item), base_currency, display_rates) / total for item in account_snapshots), default=0)


def _account_to_dict(account, snapshot, display_currency: str = "", display_rates: dict[str, float] | None = None) -> dict:
    currency = (_snapshot_currency(snapshot) if snapshot else "") or getattr(account, "base_currency", "")
    display_currency = display_currency or currency
    total_assets = snapshot.total_assets if snapshot else getattr(account, "total_assets", 0)
    cash = snapshot.cash if snapshot else getattr(account, "cash", 0)
    market_value = snapshot.market_value if snapshot else getattr(account, "market_value", 0)
    display_total = _money_to_base(total_assets, currency, display_currency, display_rates) if display_currency else total_assets
    return {
        "account_id": snapshot.account_id if snapshot else getattr(account, "account_id", ""),
        "source_name": getattr(account, "source_name", "futu") if account else "unknown",
        "broker_provider": _account_broker_provider(account) if account else "",
        "display_name": getattr(account, "display_name", "") if account else getattr(snapshot, "account_id", ""),
        "institution": getattr(account, "institution", "") if account else "",
        "import_mode": getattr(account, "import_mode", "") if account else "",
        "account_type": getattr(account, "account_type", "") if account else "",
        "markets": getattr(account, "markets", []) if account else [],
        "enabled": getattr(account, "enabled", True) if account else True,
        "base_currency": currency,
        "total_assets": total_assets,
        "cash": cash,
        "market_value": market_value,
        "display_currency": display_currency,
        "display_total_assets": display_total,
        "last_sync_time": getattr(account, "last_sync_time", None) if account else None,
        "snapshot_time": snapshot.snapshot_time if snapshot else None,
        "import_modes": _account_import_modes(getattr(account, "import_mode", "") if account else "", _account_broker_provider(account) if account else ""),
        "position_import_modes": _account_specific_import_modes(account, "position_import_modes"),
        "review_import_modes": _account_specific_import_modes(account, "review_import_modes"),
        "market_data_provider": getattr(account, "market_data_provider", "") if account else "",
        "news_data_provider": getattr(account, "news_data_provider", "") if account else "",
    }


def _account_import_modes(import_mode: str, broker_provider: str) -> list[str]:
    raw = set(_normalize_import_modes(import_mode))
    modes: list[str] = []
    if "api" in raw or broker_provider:
        modes.append("api")
    if "local" in raw or not broker_provider:
        modes.append("local")
    return modes or ["local"]


def _account_specific_import_modes(account, field_name: str) -> list[str]:
    if not account:
        return ["local"]
    raw = str(getattr(account, field_name, "") or "").strip()
    normalized = _normalize_import_modes(raw)
    if normalized:
        return normalized
    return _account_import_modes(str(getattr(account, "import_mode", "") or "").strip(), _account_broker_provider(account))


def _account_broker_provider(account) -> str:
    if not account:
        return ""
    explicit = str(getattr(account, "broker_provider", "") or "").strip().lower()
    if explicit:
        return explicit
    # Legacy fallback for local databases created before broker_provider existed.
    # Do not use source_name here; it is a user/import label and may be arbitrary.
    text = " ".join(
        str(getattr(account, key, "") or "").lower()
        for key in ("institution", "display_name")
    )
    import_mode = ",".join(_normalize_import_modes(getattr(account, "import_mode", "")))
    if "api" in import_mode and ("futu" in text or "moomoo" in text):
        return "futu"
    return ""


def _account_data_provider_preference(account: Account, data_type: str) -> str:
    if data_type in {"quote", "kline", "technical"}:
        return str(getattr(account, "market_data_provider", "") or "").strip().lower()
    if data_type in {"news", "announcement", "filing"}:
        return str(getattr(account, "news_data_provider", "") or "").strip().lower()
    return ""


def _provider_supports(provider, data_type: str, market: str) -> bool:
    normalized_market = normalize_market(market)
    return any(cap.data_type == data_type and normalized_market in cap.markets for cap in provider.capabilities())


def _normalize_import_modes(value) -> list[str]:
    if isinstance(value, list):
        raw_items = value
    else:
        raw_items = str(value or "").split(",")
    aliases = {
        "api": "api",
        "api导入": "api",
        "api 导入": "api",
        "local": "local",
        "本地导入": "local",
        "manual": "local",
        "excel": "local",
        "file": "local",
        "pdf": "local",
        "screenshot": "local",
    }
    modes: list[str] = []
    for item in raw_items:
        mode = aliases.get(str(item).strip().lower())
        if mode and mode not in modes:
            modes.append(mode)
    return modes


def _display_rate_targets(base_currency: str, account_snapshots, positions) -> list[str]:
    currencies = [base_currency, "CNY", "USD", "HKD", "CNH"]
    for snapshot in account_snapshots:
        currencies.append(_snapshot_currency(snapshot))
    for position in positions:
        currencies.extend([position.raw_currency, position.normalized_currency])
    preferred_order = [base_currency, "CNY", "USD", "HKD", "CNH"]
    result: list[str] = []
    for currency in [*preferred_order, *currencies]:
        text = str(currency or "").strip().upper()
        if text and text not in result:
            result.append(text)
    return result


def _decision_card_state(positions, cards) -> dict:
    card_by_code = {card.code: card for card in cards}
    missing = [position.code for position in positions if position.code not in card_by_code]
    stale = [
        position.code
        for position in positions
        if position.code in card_by_code and position.snapshot_time > card_by_code[position.code].data_time
    ]
    legacy = [card.code for card in cards if card.generation_source != "ai"]
    return {
        "total_positions": len(positions),
        "card_count": len(card_by_code),
        "missing_codes": missing,
        "stale_codes": stale,
        "legacy_codes": legacy,
        "needs_generation": bool(missing or stale or legacy),
    }


def _opend_health_with_timeout() -> tuple[bool, str]:
    settings = get_settings()
    try:
        with socket.create_connection((settings.futu_host, settings.futu_port), timeout=0.4):
            return True, "connected"
    except OSError as exc:
        return False, f"OpenD 未连接：{str(exc)[:80]}"
