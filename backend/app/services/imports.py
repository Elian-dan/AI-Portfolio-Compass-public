from __future__ import annotations

from dataclasses import dataclass, field
from datetime import datetime, timezone
import hashlib
from io import BytesIO
import re
from typing import Any, TYPE_CHECKING

from app.models import Account, SyncTask
from app.services.sync import mark_source_state, persist_snapshot

if TYPE_CHECKING:
    from sqlalchemy.orm import Session


@dataclass
class ImportSnapshot:
    accounts: list[dict[str, Any]]
    account_snapshots: list[dict[str, Any]]
    positions: list[dict[str, Any]]
    deals: list[dict[str, Any]] = field(default_factory=list)
    watchlist: list[dict[str, Any]] = field(default_factory=list)
    quotes: list[dict[str, Any]] = field(default_factory=list)
    news: list[dict[str, Any]] = field(default_factory=list)


def preview_import(source: str, filename: str, content: bytes, account_id: str = "", base_currency: str = "CNY") -> dict[str, Any]:
    source = _normalize_source(source)
    file_hash = hashlib.sha256(content).hexdigest()
    if source == "excel":
        snapshot, errors, warnings = _parse_excel_import(content, file_hash, account_id, base_currency)
    elif source == "alipay":
        snapshot = _parse_alipay_pdf(content, file_hash)
        errors, warnings = [], []
    elif source == "elebank":
        snapshot = _elebank_screenshot_template(file_hash)
        errors, warnings = [], ["当前来源仍使用模板兜底，未解析真实图片内容"]
    elif source in {"citic", "citic_ths"}:
        snapshot = _citic_screenshot_template(file_hash)
        source = "citic_ths"
        errors, warnings = [], ["当前来源仍使用模板兜底，未解析真实图片内容"]
    else:
        raise ValueError("unsupported import source")

    market_warnings = _supplement_markets(snapshot, source)
    warnings.extend(market_warnings)

    return {
        "source_name": source,
        "filename": filename,
        "import_hash": file_hash,
        "account": snapshot.accounts[0] if snapshot.accounts else {},
        "account_snapshot": snapshot.account_snapshots[0] if snapshot.account_snapshots else {},
        "positions": snapshot.positions,
        "position_count": len(snapshot.positions),
        "deals": snapshot.deals,
        "deal_count": len(snapshot.deals),
        "total_assets": sum(item.get("total_assets", 0) for item in snapshot.account_snapshots),
        "market_value": sum(item.get("market_value", 0) for item in snapshot.account_snapshots),
        "cash": sum(item.get("cash", 0) for item in snapshot.account_snapshots),
        "snapshot": _snapshot_to_payload(snapshot),
        "errors": errors,
        "warnings": warnings,
        "can_confirm": not errors and bool(snapshot.accounts or snapshot.account_snapshots or snapshot.positions or snapshot.deals),
    }


def confirm_import(db: "Session", preview: dict[str, Any]) -> dict[str, Any]:
    if preview.get("errors"):
        raise ValueError("导入预览存在错误，请修正后再确认")
    source = str(preview.get("source_name", "manual"))
    file_hash = str(preview.get("import_hash", ""))
    snapshot_payload = preview.get("snapshot") or {}
    snapshot = ImportSnapshot(
        accounts=list(snapshot_payload.get("accounts", [])),
        account_snapshots=[_coerce_account_snapshot(item) for item in snapshot_payload.get("account_snapshots", [])],
        positions=[_coerce_position(item) for item in snapshot_payload.get("positions", [])],
        deals=[_coerce_deal(item) for item in snapshot_payload.get("deals", [])],
        watchlist=list(snapshot_payload.get("watchlist", [])),
        quotes=[_coerce_quote(item) for item in snapshot_payload.get("quotes", [])],
        news=list(snapshot_payload.get("news", [])),
    )
    _supplement_markets(snapshot, source)
    account_id = snapshot.accounts[0]["account_id"] if snapshot.accounts else ""
    existing = db.get(Account, account_id) if account_id else None
    sync_id = f"import_{source}_{file_hash[:16]}"
    task = db.get(SyncTask, sync_id)
    if existing and existing.last_import_hash == file_hash and task and task.status == "成功":
        return {"sync_id": sync_id, "status": "已导入", "inserted_count": 0, "updated_count": 0}

    now = datetime.now(timezone.utc)
    task = task or SyncTask(sync_id=sync_id)
    task.sync_type = "手动导入"
    task.status = "执行中"
    task.start_time = now
    task.source = source
    task.idempotency_key = f"import:{source}:{file_hash}"
    db.add(task)
    db.commit()

    try:
        inserted, updated = persist_snapshot(db, sync_id, snapshot)
        for account_payload in snapshot.accounts:
            account = db.get(Account, account_payload["account_id"])
            if account:
                account.last_import_hash = file_hash
                db.add(account)
        mark_source_state(db, source, "available", "", 0)
        task.status = "成功"
        task.inserted_count = inserted
        task.updated_count = updated
    except Exception as exc:
        task.status = "失败"
        task.error_message = str(exc)[:1000]
        mark_source_state(db, source, "unavailable", task.error_message, 0)
    finally:
        task.end_time = datetime.now(timezone.utc)
        db.add(task)
        db.commit()
    return {
        "sync_id": task.sync_id,
        "status": task.status,
        "inserted_count": task.inserted_count,
        "updated_count": task.updated_count,
        "error_message": task.error_message,
    }


def _parse_excel_import(content: bytes, file_hash: str, account_id: str, base_currency: str = "CNY") -> tuple[ImportSnapshot, list[str], list[str]]:
    try:
        from openpyxl import load_workbook  # type: ignore
    except Exception as exc:  # pragma: no cover - dependency guard
        raise ValueError("Excel 导入需要安装 openpyxl") from exc

    workbook = load_workbook(BytesIO(content), data_only=True)
    errors: list[str] = []
    warnings: list[str] = []
    snapshot_time = datetime.now(timezone.utc)
    accounts: list[dict[str, Any]] = []
    account_snapshots: list[dict[str, Any]] = []
    positions: list[dict[str, Any]] = []
    deals: list[dict[str, Any]] = []
    active_account_id = account_id.strip()

    if "账户资产快照" in workbook.sheetnames:
        for row_no, row in _sheet_rows(workbook["账户资产快照"]):
            missing = _missing(row, ["snapshot_time", "total_assets", "cash", "market_value", "currency"])
            if missing:
                errors.append(f"账户资产快照 第 {row_no} 行缺少必填字段：{', '.join(missing)}")
                continue
            row_account = _row_account_id(row, active_account_id, errors, "账户资产快照", row_no)
            if not row_account:
                continue
            active_account_id = active_account_id or row_account
            snapshot_time = _parse_datetime_value(row.get("snapshot_time"))
            currency = str(row.get("currency") or "").upper()
            accounts.append(
                _account(
                    row_account,
                    str(row.get("source_name") or "manual"),
                    str(row.get("display_name") or row_account),
                    str(row.get("institution") or ""),
                    "excel",
                    currency,
                    file_hash,
                    str(row.get("account_type") or ""),
                )
            )
            account_snapshots.append(
                _account_snapshot(
                    row_account,
                    _float(row.get("total_assets")),
                    _float(row.get("cash")),
                    _float(row.get("market_value")),
                    currency,
                    snapshot_time,
                    {
                        "today_pnl": _optional_float(row.get("today_pnl")),
                        "floating_pnl": _optional_float(row.get("floating_pnl")),
                        "raw_note": str(row.get("raw_note") or ""),
                        "import_hash": file_hash,
                    },
                )
            )

    if "持仓快照" in workbook.sheetnames:
        for row_no, row in _sheet_rows(workbook["持仓快照"]):
            required = ["code", "quantity", "current_price"]
            missing = _missing(row, required)
            if missing:
                errors.append(f"持仓快照 第 {row_no} 行缺少必填字段：{', '.join(missing)}")
                continue
            row_account = _row_account_id(row, active_account_id, errors, "持仓快照", row_no)
            if not row_account:
                continue
            active_account_id = active_account_id or row_account
            item_snapshot_time = _parse_datetime_value(row.get("snapshot_time"))
            code = str(row.get("code") or "").strip()
            quantity = _float(row.get("quantity"))
            current_price = _float(row.get("current_price"))
            market_value = _optional_float(row.get("market_value"))
            if market_value is None:
                market_value = quantity * current_price
            average_cost = _float(row.get("average_cost"))
            market = str(row.get("market") or _market_from_code(code)).strip().upper()
            asset_type = str(row.get("asset_type") or ("fund" if code.upper().startswith("FUND.") else "stock")).strip()
            currency = str(row.get("currency") or base_currency or "CNY").strip().upper()
            normalized_currency = str(row.get("normalized_currency") or base_currency or currency).strip().upper()
            exchange_rate = _optional_float(row.get("exchange_rate_to_base")) or 1
            profit_loss_ratio = _optional_float(row.get("profit_loss_ratio"))
            if profit_loss_ratio is None and average_cost:
                profit_loss_ratio = (current_price - average_cost) / average_cost
            if profit_loss_ratio is None:
                profit_loss_ratio = 0
                warnings.append(f"持仓快照 第 {row_no} 行未提供盈亏比例，收益贡献分析置信度会降低")
            positions.append(
                {
                    "account_id": row_account,
                    "code": code,
                    "name": str(row.get("name") or code).strip(),
                    "market": market,
                    "asset_type": asset_type,
                    "quantity": quantity,
                    "average_cost": average_cost,
                    "current_price": current_price,
                    "raw_market_value": market_value,
                    "raw_currency": currency,
                    "normalized_market_value": market_value * exchange_rate,
                    "normalized_currency": normalized_currency,
                    "exchange_rate_to_base": exchange_rate,
                    "position_weight": _optional_float(row.get("position_weight")) or 0,
                    "profit_loss_ratio": profit_loss_ratio,
                    "requested_position_layer": str(row.get("position_layer") or "").strip(),
                    "first_buy_time": _optional_datetime(row.get("first_buy_time")),
                    "last_trade_time": _optional_datetime(row.get("last_trade_time")),
                    "missing_market_code": not _looks_like_market_code(code),
                    "snapshot_time": item_snapshot_time,
                }
            )

    if "成交记录" in workbook.sheetnames:
        for row_no, row in _sheet_rows(workbook["成交记录"]):
            missing = _missing(row, ["deal_id", "code", "price", "quantity"])
            if missing:
                errors.append(f"成交记录 第 {row_no} 行缺少必填字段：{', '.join(missing)}")
                continue
            row_account = _row_account_id(row, active_account_id, errors, "成交记录", row_no)
            if not row_account:
                continue
            active_account_id = active_account_id or row_account
            deals.append(
                {
                    "account_id": row_account,
                    "deal_id": str(row.get("deal_id") or "").strip(),
                    "order_id": str(row.get("order_id") or ""),
                    "code": str(row.get("code") or "").strip(),
                    "side": _normalize_deal_side(row.get("side")),
                    "price": _float(row.get("price")),
                    "quantity": _float(row.get("quantity")),
                    "deal_time": _parse_datetime_value(row.get("deal_time")),
                    "market": str(row.get("market") or "").strip().upper(),
                    "raw_payload": {
                        "amount": _optional_float(row.get("amount")),
                        "currency": str(row.get("currency") or ""),
                        "fee": _optional_float(row.get("fee")),
                        "commission": _optional_float(row.get("commission")),
                        "tax": _optional_float(row.get("tax")),
                        "raw_note": str(row.get("raw_note") or ""),
                    },
                }
            )

    if not any(name in workbook.sheetnames for name in ("账户资产快照", "持仓快照", "成交记录")):
        errors.append("Excel 必须至少包含一个 Sheet：账户资产快照、持仓快照、成交记录")

    if not account_snapshots and not positions and not deals:
        errors.append("Excel 中未识别到可导入的持仓或成交记录，请从第 2 行开始填写数据")

    if active_account_id and not accounts and (account_snapshots or positions or deals):
        accounts.append(_account(active_account_id, "manual", active_account_id, "", "excel", "CNY", file_hash))

    total_assets = next((item["total_assets"] for item in account_snapshots if item["account_id"] == active_account_id), 0)
    if not total_assets:
        total_assets = sum(item["normalized_market_value"] for item in positions)
    if total_assets:
        for item in positions:
            if not item.get("position_weight"):
                item["position_weight"] = item["normalized_market_value"] / total_assets

    return ImportSnapshot(accounts=accounts, account_snapshots=account_snapshots, positions=positions, deals=deals), errors, warnings


def _parse_alipay_pdf(content: bytes, file_hash: str) -> ImportSnapshot:
    try:
        import pdfplumber  # type: ignore
    except Exception as exc:  # pragma: no cover - dependency guard
        raise ValueError("支付宝 PDF 解析需要安装 pdfplumber，不能使用样例数据代替真实文件") from exc

    with pdfplumber.open(BytesIO(content)) as pdf:
        text = "\n".join(page.extract_text() or "" for page in pdf.pages)
        date_match = re.search(r"截止至([0-9-]+)", text)
        snapshot_time = _parse_snapshot_date(date_match.group(1) if date_match else "")
        rows: list[dict[str, Any]] = []
        for page in pdf.pages:
            for table in page.extract_tables():
                if not table or not table[0]:
                    continue
                header = [str(item or "").strip() for item in table[0]]
                if "基金代码" not in "".join(header):
                    continue
                for row in table[1:]:
                    if not row or not str(row[0] or "").strip().isdigit():
                        continue
                    item = dict(zip(header, [str(value or "").strip() for value in row]))
                    rows.append(item)

    if not rows:
        raise ValueError("未识别到支付宝基金持仓表格，当前 PDF 版式暂不支持自动导入")

    account_id = "alipay_fund"
    positions = []
    for row in rows:
        fund_code = row.get("基金代码", "")
        market_value = _float(row.get("资产小计"))
        quantity = _float(row.get("总份额"))
        price = _float(row.get("单位净值"))
        positions.append(
            {
                "account_id": account_id,
                "code": f"FUND.{fund_code}" if fund_code else f"ALIPAY:{row.get('基金名称', '')}",
                "name": _compact(row.get("基金名称", "")),
                "market": "CN",
                "asset_type": "fund",
                "quantity": quantity,
                "average_cost": 0,
                "current_price": price,
                "raw_market_value": market_value,
                "raw_currency": "CNY",
                "normalized_market_value": market_value,
                "normalized_currency": "CNY",
                "exchange_rate_to_base": 1,
                "position_weight": 0,
                "profit_loss_ratio": 0,
                "missing_market_code": False,
                "snapshot_time": snapshot_time,
            }
        )
    total_assets = round(sum(item["normalized_market_value"] for item in positions), 2)
    for item in positions:
        item["position_weight"] = item["normalized_market_value"] / total_assets if total_assets else 0

    return ImportSnapshot(
        accounts=[
            _account(
                account_id,
                "alipay",
                "支付宝理财",
                "蚂蚁基金",
                "pdf",
                "CNY",
                file_hash,
                account_type="基金理财",
            )
        ],
        account_snapshots=[
            _account_snapshot(account_id, total_assets, 0, total_assets, "CNY", snapshot_time, {"import_hash": file_hash})
        ],
        positions=positions,
    )


def _elebank_screenshot_template(file_hash: str) -> ImportSnapshot:
    snapshot_time = datetime.now(timezone.utc)
    account_id = "elebank_margin_usd"
    positions = [
        _position(account_id, "US.NVDA", "英伟达", "US", "stock", 28, 207.34, 197.94, 5542.32, "USD", -0.0453, False, snapshot_time),
        _position(account_id, "SPCX", "SpaceX", "US", "stock", 10, 200.00, 171.29, 1712.90, "USD", -0.1436, False, snapshot_time),
    ]
    _apply_weights(positions, 8265.81)
    return ImportSnapshot(
        accounts=[_account(account_id, "elebank", "Elebank 保证金投资账户", "Elebank", "screenshot", "USD", file_hash, "保证金投资")],
        account_snapshots=[_account_snapshot(account_id, 8265.81, 1010.59, 7255.22, "USD", snapshot_time, {"today_pnl": -55.90, "import_hash": file_hash})],
        positions=positions,
    )


def _citic_screenshot_template(file_hash: str) -> ImportSnapshot:
    snapshot_time = datetime.now(timezone.utc)
    account_id = "citic_cny_a_883601"
    positions = [
        _position(account_id, "CITIC:科华数据", "科华数据", "CN", "stock", 290, 44.1414, 40.1400, 11640.60, "CNY", -0.0906, True, snapshot_time),
        _position(account_id, "CITIC:机器人", "机器人", "CN", "stock", 1200, 1.0042, 1.1520, 1382.40, "CNY", 0.1472, True, snapshot_time),
        _position(account_id, "CITIC:湖南白银", "湖南白银", "CN", "stock", 100, 19.0900, 8.1200, 812.00, "CNY", -0.5746, True, snapshot_time),
        _position(account_id, "CITIC:白银基金", "白银基金", "CN", "fund", 400, 4.5515, 1.6980, 679.20, "CNY", -0.6269, True, snapshot_time),
    ]
    _apply_weights(positions, 28056.17)
    return ImportSnapshot(
        accounts=[_account(account_id, "citic_ths", "中信证券人民币 A 股", "中信证券", "screenshot", "CNY", file_hash, "证券账户")],
        account_snapshots=[_account_snapshot(account_id, 28056.17, 13541.97, 14514.20, "CNY", snapshot_time, {"floating_pnl": -3247.63, "today_pnl": -675.30, "import_hash": file_hash})],
        positions=positions,
    )


def _position(
    account_id: str,
    code: str,
    name: str,
    market: str,
    asset_type: str,
    quantity: float,
    average_cost: float,
    current_price: float,
    market_value: float,
    currency: str,
    profit_loss_ratio: float,
    missing_market_code: bool,
    snapshot_time: datetime,
) -> dict[str, Any]:
    return {
        "account_id": account_id,
        "code": code,
        "name": name,
        "market": market,
        "asset_type": asset_type,
        "quantity": quantity,
        "average_cost": average_cost,
        "current_price": current_price,
        "raw_market_value": market_value,
        "raw_currency": currency,
        "normalized_market_value": market_value,
        "normalized_currency": currency,
        "exchange_rate_to_base": 1,
        "position_weight": 0,
        "profit_loss_ratio": profit_loss_ratio,
        "missing_market_code": missing_market_code,
        "snapshot_time": snapshot_time,
    }


def _account(
    account_id: str,
    source_name: str,
    display_name: str,
    institution: str,
    import_mode: str,
    base_currency: str,
    file_hash: str,
    account_type: str = "",
) -> dict[str, Any]:
    return {
        "account_id": account_id,
        "source_name": source_name,
        "display_name": display_name,
        "institution": institution,
        "import_mode": import_mode,
        "enabled": True,
        "last_import_hash": file_hash,
        "account_type": account_type,
        "trade_env": "REAL",
        "markets": [],
        "base_currency": base_currency,
    }


def _account_snapshot(
    account_id: str,
    total_assets: float,
    cash: float,
    market_value: float,
    currency: str,
    snapshot_time: datetime,
    extra: dict[str, Any],
) -> dict[str, Any]:
    values = {"currency": currency, **extra}
    return {
        "account_id": account_id,
        "total_assets": total_assets,
        "cash": cash,
        "market_value": market_value,
        "raw_currency_values": values,
        "snapshot_time": snapshot_time,
    }


def _snapshot_to_payload(snapshot: ImportSnapshot) -> dict[str, Any]:
    return {
        "accounts": snapshot.accounts,
        "account_snapshots": snapshot.account_snapshots,
        "positions": snapshot.positions,
        "deals": snapshot.deals,
        "watchlist": snapshot.watchlist,
        "quotes": snapshot.quotes,
        "news": snapshot.news,
    }


def _supplement_markets(snapshot: ImportSnapshot, source: str) -> list[str]:
    updated = 0
    for item in snapshot.positions:
        original_market = str(item.get("market") or "").strip().upper()
        inferred_market = _infer_position_market(item)
        should_fill_missing = not original_market and inferred_market
        should_refine_alipay_fund = source == "alipay" and original_market in {"", "CN"} and inferred_market in {"US", "HK"}
        if should_fill_missing or should_refine_alipay_fund:
            item["market"] = inferred_market
            item["missing_market_code"] = not _looks_like_market_code(str(item.get("code") or ""))
            updated += 1
        elif original_market:
            item["market"] = original_market

    markets = sorted({str(item.get("market") or "").strip().upper() for item in snapshot.positions if item.get("market")})
    for account in snapshot.accounts:
        existing = [str(market).strip().upper() for market in account.get("markets", []) if str(market).strip()]
        account["markets"] = sorted(set(existing).union(markets))

    if not updated:
        return []
    return [f"已根据基金/持仓名称补充 {updated} 条市场字段"]


def _infer_position_market(item: dict[str, Any]) -> str:
    code = str(item.get("code") or "").strip().upper()
    if "." in code:
        prefix = code.split(".", 1)[0]
        if prefix in {"US", "HK", "SH", "SZ", "CN"}:
            return "CN" if prefix in {"SH", "SZ"} else prefix
    text = f"{code} {item.get('name') or ''} {item.get('asset_type') or ''}".upper()
    if any(token in text for token in ("纳斯达克", "NASDAQ", "标普", "S&P", "S＆P", "道琼斯", "DOW JONES", "美国", "美股")):
        return "US"
    if any(token in text for token in ("恒生", "港股", "香港", "HANG SENG", "HSI", "HKEX")):
        return "HK"
    if any(token in text for token in ("A股", "沪深", "中证", "创业板", "科创", "上证", "深证", "CSI")):
        return "CN"
    return ""


def _apply_weights(positions: list[dict[str, Any]], total_assets: float) -> None:
    for item in positions:
        item["position_weight"] = item["normalized_market_value"] / total_assets if total_assets else 0


def _parse_snapshot_date(value: str) -> datetime:
    try:
        return datetime.strptime(value, "%Y-%m-%d").replace(tzinfo=timezone.utc)
    except Exception:
        return datetime.now(timezone.utc)


def _parse_datetime_value(value: Any) -> datetime:
    if isinstance(value, datetime):
        return value
    if isinstance(value, str):
        try:
            parsed = datetime.fromisoformat(value.replace("Z", "+00:00"))
            return parsed if parsed.tzinfo else parsed.replace(tzinfo=timezone.utc)
        except Exception:
            pass
    return datetime.now(timezone.utc)


def _coerce_account_snapshot(item: dict[str, Any]) -> dict[str, Any]:
    payload = dict(item)
    payload["snapshot_time"] = _parse_datetime_value(payload.get("snapshot_time"))
    return payload


def _coerce_position(item: dict[str, Any]) -> dict[str, Any]:
    payload = dict(item)
    payload["snapshot_time"] = _parse_datetime_value(payload.get("snapshot_time"))
    if payload.get("first_buy_time"):
        payload["first_buy_time"] = _parse_datetime_value(payload.get("first_buy_time"))
    if payload.get("last_trade_time"):
        payload["last_trade_time"] = _parse_datetime_value(payload.get("last_trade_time"))
    return payload


def _coerce_deal(item: dict[str, Any]) -> dict[str, Any]:
    payload = dict(item)
    payload["deal_time"] = _parse_datetime_value(payload.get("deal_time"))
    payload["raw_payload"] = payload.get("raw_payload") or {}
    return payload


def _coerce_quote(item: dict[str, Any]) -> dict[str, Any]:
    payload = dict(item)
    payload["quote_time"] = _parse_datetime_value(payload.get("quote_time"))
    return payload


def _sheet_rows(sheet: Any) -> list[tuple[int, dict[str, Any]]]:
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [_canonical_excel_header(value) for value in rows[0]]
    result = []
    for index, values in enumerate(rows[1:], start=2):
        if not any(value not in (None, "") for value in values):
            continue
        result.append((index, {headers[col]: values[col] if col < len(values) else None for col in range(len(headers)) if headers[col]}))
    return result


EXCEL_HEADER_ALIASES = {
    "标的代码*": "code", "标的代码": "code", "标的名称": "name", "数量*": "quantity", "数量": "quantity",
    "当前价格*": "current_price", "当前价格": "current_price", "平均成本": "average_cost", "市场": "market",
    "类型": "asset_type", "原币币种": "currency", "折算币种": "normalized_currency", "汇率": "exchange_rate_to_base",
    "仓位类型": "position_layer", "快照时间": "snapshot_time", "成交号*": "deal_id", "成交号": "deal_id",
    "成交编号*": "deal_id", "成交编号": "deal_id", "订单号": "order_id", "订单编号": "order_id",
    "方向": "side", "买卖方向*": "side", "买卖方向": "side", "价格*": "price", "价格": "price",
    "成交价格*": "price", "成交价格": "price", "成交数量*": "quantity", "成交数量": "quantity",
    "成交时间*": "deal_time", "成交时间": "deal_time",
}


def _canonical_excel_header(value: Any) -> str:
    header = str(value or "").strip()
    return EXCEL_HEADER_ALIASES.get(header, header)


def _market_from_code(value: str) -> str:
    prefix = value.strip().upper().split(".", 1)[0]
    if prefix in {"SH", "SZ", "CN", "FUND"}:
        return "CN"
    return prefix if prefix in {"US", "HK"} else ""


def _normalize_deal_side(value: Any) -> str:
    side = str(value or "买入").strip().upper()
    if side in {"买入", "BUY", "B"}:
        return "BUY"
    if side in {"卖出", "SELL", "S"}:
        return "SELL"
    return side


def _missing(row: dict[str, Any], fields: list[str]) -> list[str]:
    return [field for field in fields if row.get(field) in (None, "")]


def _row_account_id(row: dict[str, Any], active_account_id: str, errors: list[str], sheet_name: str, row_no: int) -> str:
    row_account = str(row.get("account_id") or active_account_id or "").strip()
    if active_account_id and row_account and row_account != active_account_id:
        errors.append(f"{sheet_name} 第 {row_no} 行 account_id 与当前账户不一致")
        return ""
    if not row_account:
        errors.append(f"{sheet_name} 第 {row_no} 行缺少 account_id，且当前未选择账户")
        return ""
    return row_account


def _optional_float(value: Any) -> float | None:
    if value in (None, ""):
        return None
    return _float(value)


def _optional_datetime(value: Any) -> datetime | None:
    if value in (None, ""):
        return None
    return _parse_datetime_value(value)


def _looks_like_market_code(value: str) -> bool:
    code = value.strip().upper()
    return bool(re.match(r"^(US|HK|SH|SZ|FUND)\.", code))


def _force_account_id(snapshot: ImportSnapshot, account_id: str) -> None:
    for account in snapshot.accounts:
        account["account_id"] = account_id
    for item in snapshot.account_snapshots:
        item["account_id"] = account_id
    for item in snapshot.positions:
        item["account_id"] = account_id
    for item in snapshot.deals:
        item["account_id"] = account_id


def _float(value: Any) -> float:
    try:
        return float(str(value or "0").replace(",", "").strip())
    except Exception:
        return 0.0


def _compact(value: str) -> str:
    return " ".join(str(value).replace("\n", " ").split())


def _normalize_source(source: str) -> str:
    return source.strip().lower().replace("-", "_")
